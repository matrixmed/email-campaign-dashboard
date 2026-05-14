import os
import sys
import re
import json
import openpyxl
from collections import defaultdict
from dotenv import load_dotenv

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
load_dotenv(os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), '.env'))

import psycopg2
from psycopg2.extras import RealDictCursor

BASE = r'C:\Users\AndrewDaly\Desktop\email-campaign-dashboard\backend\Lists\Print Lists'
HT_BASE = os.path.join(BASE, 'Hot Topics List Changes')


def norm_str(s):
    if s is None:
        return ''
    if isinstance(s, float):
        if s.is_integer():
            s = str(int(s))
        else:
            s = str(s)
    s = str(s).strip().upper()
    s = re.sub(r'\s+', ' ', s)
    return s


def norm_zip(s):
    if s is None:
        return ''
    if isinstance(s, float):
        if s.is_integer():
            s = str(int(s))
        else:
            s = str(s)
    s = re.sub(r'\D', '', str(s))
    if len(s) > 5:
        s = s[:5]
    return s.zfill(5) if s else ''


def norm_addr(s):
    if not s:
        return ''
    s = norm_str(s)
    repl = {
        'STREET': 'ST', 'AVENUE': 'AVE', 'BOULEVARD': 'BLVD', 'DRIVE': 'DR',
        'LANE': 'LN', 'ROAD': 'RD', 'COURT': 'CT', 'PLACE': 'PL',
        'SUITE': 'STE', 'APARTMENT': 'APT', 'BUILDING': 'BLDG',
        'NORTH': 'N', 'SOUTH': 'S', 'EAST': 'E', 'WEST': 'W',
        'HIGHWAY': 'HWY', 'PARKWAY': 'PKWY', 'CIRCLE': 'CIR',
    }
    for full, abbr in repl.items():
        s = re.sub(r'\b' + full + r'\b', abbr, s)
    s = re.sub(r'[.,#]', '', s)
    s = re.sub(r'\s+', ' ', s)
    return s.strip()


def norm_name(s):
    if not s:
        return ''
    s = norm_str(s)
    s = re.sub(r'[^A-Z\s-]', '', s)
    return s.strip()


def get_npi(row):
    npi = row.get('NPI Number') or row.get('NPI') or row.get('npi')
    if npi is None:
        return None
    if isinstance(npi, float):
        if npi.is_integer():
            npi = str(int(npi))
        else:
            return None
    npi = str(npi).strip()
    if re.match(r'^\d{10}$', npi):
        return npi
    return None


def get_addr(row):
    return row.get('Address 1') or row.get('Address Line 1') or row.get('Address  1') or ''


def get_zip(row):
    return row.get('Zip') or row.get('Zip Code') or ''


def load_xlsx_rows(path, sheet=None):
    wb = openpyxl.load_workbook(path, data_only=True)
    sheets = [sheet] if sheet else wb.sheetnames
    out = []
    for sn in sheets:
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = [str(h).strip() if h else '' for h in rows[0]]
        for r in rows[1:]:
            if r is None or all(v is None or (isinstance(v, str) and not v.strip()) for v in r):
                continue
            d = {header[i]: r[i] for i in range(min(len(header), len(r)))}
            d['__sheet'] = sn
            out.append(d)
    wb.close()
    return out


def excel_key(row):
    npi = get_npi(row)
    if npi:
        return ('npi', npi)
    fn = norm_name(row.get('First Name', ''))
    ln = norm_name(row.get('Last Name', ''))
    addr = norm_addr(get_addr(row))
    city = norm_str(row.get('City', ''))
    state = norm_str(row.get('State', ''))
    zc = norm_zip(get_zip(row))
    return ('nm', fn, ln, addr, city, state, zc)


def db_key(row):
    npi = (row.get('npi') or '').strip()
    if npi and re.match(r'^\d{10}$', npi):
        return ('npi', npi)
    return (
        'nm',
        norm_name(row.get('first_name')),
        norm_name(row.get('last_name')),
        norm_addr(row.get('address_1')),
        norm_str(row.get('city')),
        norm_str(row.get('state')),
        norm_zip(row.get('zipcode')),
    )


def fetch_subs(conn, list_name):
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute("""
        SELECT id, npi, first_name, last_name, degree, specialty, type_of_professional,
               company, title, address_1, address_2, city, state, zipcode,
               subscribed_lists, unsubscribed_lists, is_comp
        FROM print_list_subscribers
        WHERE subscribed_lists LIKE %s AND is_subscribed = TRUE
    """, (f'%{list_name}%',))
    rows = cur.fetchall()
    cur.close()
    return [r for r in rows if list_name in (r.get('subscribed_lists') or '').split(',')]


def diff_print_list(conn, list_name, print_file, comp_file=None, comp_sheet=None, label=None):
    label = label or list_name
    print_rows = load_xlsx_rows(os.path.join(BASE, print_file))
    for r in print_rows:
        r['__is_comp'] = False
        r['__source'] = 'print'
    comp_rows = []
    if comp_file:
        comp_rows = load_xlsx_rows(os.path.join(BASE, comp_file), sheet=comp_sheet)
        for r in comp_rows:
            r['__is_comp'] = True
            r['__source'] = 'comp'

    excel_rows = print_rows + comp_rows
    db_rows = fetch_subs(conn, list_name)

    excel_by_key = {}
    excel_by_name = defaultdict(list)
    dupe_excel_keys = 0
    for r in excel_rows:
        k = excel_key(r)
        if k in excel_by_key:
            dupe_excel_keys += 1
            continue
        excel_by_key[k] = r
        nk = (norm_name(r.get('First Name', '')), norm_name(r.get('Last Name', '')))
        excel_by_name[nk].append(r)

    db_by_key = {}
    db_by_name = defaultdict(list)
    dupe_db_keys = 0
    for r in db_rows:
        k = db_key(r)
        if k in db_by_key:
            dupe_db_keys += 1
            continue
        db_by_key[k] = r
        nk = (norm_name(r.get('first_name')), norm_name(r.get('last_name')))
        db_by_name[nk].append(r)

    adds = []
    edits = []
    flag_flips = []
    matched_db_ids = set()

    for k, er in excel_by_key.items():
        match = db_by_key.get(k)
        if not match:
            nk = (norm_name(er.get('First Name', '')), norm_name(er.get('Last Name', '')))
            cands = db_by_name.get(nk, [])
            for cand in cands:
                if cand['id'] in matched_db_ids:
                    continue
                e_addr = norm_addr(get_addr(er))
                d_addr = norm_addr(cand.get('address_1'))
                if e_addr and d_addr and (e_addr[:10] == d_addr[:10] or norm_str(er.get('City')) == norm_str(cand.get('city'))):
                    match = cand
                    break
        if not match:
            adds.append(er)
            continue
        matched_db_ids.add(match['id'])

        e_addr = norm_addr(get_addr(er))
        e_city = norm_str(er.get('City'))
        e_state = norm_str(er.get('State'))
        e_zip = norm_zip(get_zip(er))
        e_fn = norm_name(er.get('First Name'))
        e_ln = norm_name(er.get('Last Name'))
        e_comp = er['__is_comp']

        d_addr = norm_addr(match.get('address_1'))
        d_city = norm_str(match.get('city'))
        d_state = norm_str(match.get('state'))
        d_zip = norm_zip(match.get('zipcode'))
        d_fn = norm_name(match.get('first_name'))
        d_ln = norm_name(match.get('last_name'))
        d_comp = bool(match.get('is_comp'))

        field_diffs = {}
        if e_addr and e_addr != d_addr:
            field_diffs['address_1'] = (match.get('address_1'), get_addr(er))
        if e_city and e_city != d_city:
            field_diffs['city'] = (match.get('city'), er.get('City'))
        if e_state and e_state != d_state:
            field_diffs['state'] = (match.get('state'), er.get('State'))
        if e_zip and d_zip and e_zip != d_zip:
            field_diffs['zipcode'] = (match.get('zipcode'), get_zip(er))
        if e_fn != d_fn and e_fn:
            field_diffs['first_name'] = (match.get('first_name'), er.get('First Name'))
        if e_ln != d_ln and e_ln:
            field_diffs['last_name'] = (match.get('last_name'), er.get('Last Name'))

        flag_changed = e_comp != d_comp

        if field_diffs:
            edits.append({'db': match, 'excel': er, 'diffs': field_diffs, 'flag_flip': flag_changed, 'new_is_comp': e_comp})
        elif flag_changed:
            flag_flips.append({'db': match, 'excel': er, 'new_is_comp': e_comp})

    removes = []
    for r in db_rows:
        if r['id'] not in matched_db_ids:
            removes.append(r)

    return {
        'label': label,
        'list_name': list_name,
        'excel_count': len(excel_rows),
        'excel_print': len(print_rows),
        'excel_comp': len(comp_rows),
        'dupe_excel_keys': dupe_excel_keys,
        'db_count': len(db_rows),
        'dupe_db_keys': dupe_db_keys,
        'adds': adds,
        'edits': edits,
        'flag_flips': flag_flips,
        'removes': removes,
    }


def diff_unsub_sheet(conn, excel_path, sheet, list_token, label=None):
    label = label or f'{os.path.basename(excel_path)}::{sheet}'
    rows = load_xlsx_rows(excel_path, sheet)
    cur = conn.cursor(cursor_factory=RealDictCursor)

    pending = []
    already = []
    not_found = []
    multi = []

    for er in rows:
        fn_raw = er.get('First Name')
        ln_raw = er.get('Last Name')
        if not fn_raw or not ln_raw:
            continue
        fn = norm_name(fn_raw)
        ln = norm_name(ln_raw)
        addr = norm_addr(get_addr(er))
        city = norm_str(er.get('City'))
        state = norm_str(er.get('State'))
        zc = norm_zip(get_zip(er))

        cur.execute("""
            SELECT id, npi, first_name, last_name, address_1, city, state, zipcode,
                   subscribed_lists, unsubscribed_lists, is_subscribed
            FROM print_list_subscribers
            WHERE UPPER(TRIM(first_name)) = %s AND UPPER(TRIM(last_name)) = %s
        """, (fn, ln))
        cands = cur.fetchall()
        if not cands:
            not_found.append({'excel': er})
            continue

        scored = []
        for c in cands:
            score = 0
            if addr and norm_addr(c.get('address_1')) == addr:
                score += 5
            elif addr and c.get('address_1') and norm_addr(c.get('address_1'))[:10] == addr[:10]:
                score += 3
            if city and norm_str(c.get('city')) == city:
                score += 2
            if state and norm_str(c.get('state')) == state:
                score += 1
            if zc and norm_zip(c.get('zipcode')) == zc:
                score += 1
            scored.append((score, c))
        scored.sort(key=lambda x: -x[0])
        best_score, best = scored[0]
        if len(scored) > 1 and scored[1][0] == best_score and best_score < 5:
            multi.append({'excel': er, 'candidates': [s[1] for s in scored if s[0] == best_score]})
            continue
        if best_score == 0:
            not_found.append({'excel': er, 'candidates': cands[:5]})
            continue

        cur_sub = (best.get('subscribed_lists') or '')
        cur_unsub = (best.get('unsubscribed_lists') or '')
        sub_tokens = [x.strip() for x in cur_sub.split(',') if x.strip()]
        unsub_tokens = [x.strip() for x in cur_unsub.split(',') if x.strip()]

        if list_token not in sub_tokens and list_token in unsub_tokens:
            already.append({'excel': er, 'db': best})
            continue
        if list_token not in sub_tokens and list_token not in unsub_tokens:
            already.append({'excel': er, 'db': best})
            continue

        pending.append({
            'excel': er,
            'db': best,
            'list_token': list_token,
            'reason': er.get('Notes') or 'Per source-of-truth file',
        })

    cur.close()
    return {
        'label': label,
        'list_token': list_token,
        'excel_count': len(rows),
        'pending': pending,
        'already': already,
        'not_found': not_found,
        'multi': multi,
    }


def summarize_list(d, name):
    print(f'\n=== {name} ===')
    print(f"  excel: {d['excel_count']} (print={d['excel_print']}, comp={d['excel_comp']})  db: {d['db_count']}  dupe_excel={d['dupe_excel_keys']}  dupe_db={d['dupe_db_keys']}")
    print(f"  adds:       {len(d['adds'])}")
    print(f"  field edits:{len(d['edits'])}")
    print(f"  flag flips: {len(d['flag_flips'])}")
    print(f"  removes:    {len(d['removes'])}")


def summarize_unsub(d, name):
    print(f'\n=== {name} ===')
    print(f"  excel rows: {d['excel_count']}")
    print(f"  pending unsubscribes: {len(d['pending'])}")
    print(f"  already unsubscribed: {len(d['already'])}")
    print(f"  not found in DB:      {len(d['not_found'])}")
    print(f"  ambiguous (multi):    {len(d['multi'])}")


def serialize_row_kv(r):
    out = {}
    for k, v in r.items():
        if k.startswith('__'):
            continue
        if hasattr(v, 'isoformat'):
            out[k] = v.isoformat()
        elif isinstance(v, (str, int, float, type(None), bool)):
            out[k] = v
        else:
            out[k] = str(v)
    return out


def main():
    conn = psycopg2.connect(os.getenv('DATABASE_URL'))
    results = {}

    print('Computing diffs (per-list combining Print + Comp)...')

    results['JCAD'] = diff_print_list(conn, 'JCAD', 'JCAD Print List.xlsx', 'JCAD Comp List.xlsx')
    summarize_list(results['JCAD'], 'JCAD (Print + Comp)')

    results['NPPA'] = diff_print_list(conn, 'NPPA', 'NP+PA Print List.xlsx')
    summarize_list(results['NPPA'], 'NPPA (Print)')

    results['BT'] = diff_print_list(conn, 'BT', 'BT Print List.xlsx', 'BT Comp List.xlsx')
    summarize_list(results['BT'], 'BT (Print + Comp)')

    print('\n--- Unsubscribe sheets ---')

    results['JCAD_Delete'] = diff_unsub_sheet(
        conn,
        os.path.join(BASE, 'JCAD List Management.xlsx'),
        'Delete',
        'JCAD',
        label='JCAD List Management::Delete',
    )
    summarize_unsub(results['JCAD_Delete'], 'JCAD List Mgmt - Delete')

    ht_files = [
        'Dermatology List Updates.xlsx',
        'Diabetes List Updates.xlsx',
        'Gastro List Updates.xlsx',
        'Immunology List Changes.xlsx',
        'Neurology List Updates.xlsx',
        'Oncology List Updates.xlsx',
    ]
    for fname in ht_files:
        path = os.path.join(HT_BASE, fname)
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        for sn in wb.sheetnames:
            tok = f'HT-{sn}'
            key = f'HT::{fname}::{sn}'
            results[key] = diff_unsub_sheet(conn, path, sn, tok, label=key)
            summarize_unsub(results[key], key)
        wb.close()

    out_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'print_list_diff_report.json')
    serializable = {}
    for k, v in results.items():
        if 'adds' in v:
            serializable[k] = {
                'label': v['label'],
                'list_name': v.get('list_name'),
                'excel_count': v['excel_count'],
                'excel_print': v['excel_print'],
                'excel_comp': v['excel_comp'],
                'dupe_excel_keys': v['dupe_excel_keys'],
                'dupe_db_keys': v['dupe_db_keys'],
                'db_count': v['db_count'],
                'adds_count': len(v['adds']),
                'edits_count': len(v['edits']),
                'flag_flips_count': len(v['flag_flips']),
                'removes_count': len(v['removes']),
                'adds': [serialize_row_kv(r) for r in v['adds']],
                'edits': [{'db_id': e['db']['id'], 'db_npi': e['db'].get('npi'), 'name': f"{e['db'].get('first_name')} {e['db'].get('last_name')}", 'diffs': e['diffs'], 'flag_flip': e.get('flag_flip'), 'new_is_comp': e.get('new_is_comp')} for e in v['edits']],
                'flag_flips': [{'db_id': f['db']['id'], 'db_npi': f['db'].get('npi'), 'name': f"{f['db'].get('first_name')} {f['db'].get('last_name')}", 'old_is_comp': bool(f['db'].get('is_comp')), 'new_is_comp': f['new_is_comp']} for f in v['flag_flips']],
                'removes': [{'id': r['id'], 'npi': r.get('npi'), 'first_name': r.get('first_name'), 'last_name': r.get('last_name'), 'address_1': r.get('address_1'), 'city': r.get('city'), 'state': r.get('state'), 'is_comp': r.get('is_comp'), 'subscribed_lists': r.get('subscribed_lists')} for r in v['removes']],
            }
        else:
            serializable[k] = {
                'label': v['label'],
                'list_token': v['list_token'],
                'excel_count': v['excel_count'],
                'pending_count': len(v['pending']),
                'already_count': len(v['already']),
                'not_found_count': len(v['not_found']),
                'multi_count': len(v['multi']),
                'pending': [{
                    'db_id': p['db']['id'],
                    'db_npi': p['db'].get('npi'),
                    'name': f"{p['db'].get('first_name')} {p['db'].get('last_name')}",
                    'db_address': p['db'].get('address_1'),
                    'db_city': p['db'].get('city'),
                    'list_token': p['list_token'],
                    'reason': p['reason'],
                    'currently_subscribed': p['db'].get('subscribed_lists'),
                    'currently_unsubscribed': p['db'].get('unsubscribed_lists'),
                } for p in v['pending']],
                'not_found_samples': [{'name': f"{n['excel'].get('First Name')} {n['excel'].get('Last Name')}", 'address': get_addr(n['excel']), 'city': n['excel'].get('City')} for n in v['not_found'][:25]],
                'multi_samples': [{'name': f"{m['excel'].get('First Name')} {m['excel'].get('Last Name')}", 'count': len(m['candidates'])} for m in v['multi'][:10]],
            }

    with open(out_path, 'w') as f:
        json.dump(serializable, f, indent=2, default=str)
    print(f'\nReport written to: {out_path}')

    conn.close()


if __name__ == '__main__':
    main()