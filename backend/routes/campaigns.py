from flask import Blueprint, request, jsonify
from sqlalchemy.orm import sessionmaker
from sqlalchemy import create_engine, text
from models import CampaignReportingMetadata, BrandEditorAgency, CMIContractValue, GCMPlacementLookup, UserProfile, UniversalProfile
from werkzeug.utils import secure_filename
import os
import json
import re
import threading
import time
import traceback
from datetime import datetime
import openpyxl
import tempfile
from sqlalchemy.exc import OperationalError

campaigns_bp = Blueprint('campaigns', __name__)

ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'png', 'jpg', 'jpeg'}

COLUMN_MAPPINGS = {
    'cmi_placement_id': ['cmi_placementid', 'CMI_PlacementID'],
    'client_placement_id': ['client_placementid', 'Client_PlacementID'],
    'client_id': ['client_id', 'Client_ID'],
    'placement_description': ['placement_description', 'Placement_Description'],
    'supplier': ['supplier', 'Supplier'],
    'brand_name': ['brand name', 'Brand_Name', 'brand_name'],
    'vehicle_name': ['Vehicle_Name', 'vehicle_name', 'tactic_name'],
    'target_list_id': ['targetlistid', 'TargetListID'],
    'campaign_name': ['campaign_name', 'Campaign_Name'],
    'contract_number': ['contractconfirmationnumber', 'contract_number'],
    'buying_channel': ['buying_channel'],
    'creative_code': ['creative_code'],
    'media_tactic_id': ['media_tactic_id', 'mediatacticid', 'Media_Tactic_ID'],
    'npi': ['npi', 'NPI', 'NPI Number', 'NPI #', 'NPI_Number', 'NPINumber', 'National Provider Identifier']
}

PHARMA_COMPANY_MAP = {
    'eli lilly': 'Lilly',
    'lilly': 'Lilly',
    'eli lilly & company': 'Lilly',
    'eli lilly and company': 'Lilly',
    'johnson & johnson': 'J&J',
    'johnson and johnson': 'J&J',
    'j&j': 'J&J',
    'jnj': 'J&J',
    'astrazeneca': 'AstraZeneca',
    'abbvie': 'Abbvie',
    'boehringer ingelheim': 'BI',
    'bi': 'BI',
    'exelixis': 'Exelixis',
    'dg': 'DG',
    'doctors group': 'DG',
    'dsi': 'DSI'
}

def get_session():
    engine = create_engine(os.getenv('DATABASE_URL'))
    Session = sessionmaker(bind=engine)
    return Session()

def allowed_file(filename, file_type):
    if not filename or '.' not in filename:
        return False
    ext = filename.rsplit('.', 1)[1].lower()
    if file_type == 'excel':
        return ext in {'xlsx', 'xls'}
    elif file_type == 'image':
        return ext in {'png', 'jpg', 'jpeg'}
    return ext in ALLOWED_EXTENSIONS

def find_column_value(row_data, field_name):
    possible_columns = COLUMN_MAPPINGS.get(field_name, [field_name])
    for col_name in possible_columns:
        for key in row_data:
            if key and str(key).lower().strip() == col_name.lower().strip():
                return row_data[key]
    return None

def get_pharma_company_for_brand(brand_name):
    try:
        session = get_session()
        brand_mapping = session.query(BrandEditorAgency).filter(
            BrandEditorAgency.brand.ilike(f'%{brand_name}%'),
            BrandEditorAgency.is_active == True
        ).first()
        session.close()
        if brand_mapping:
            return brand_mapping.pharma_company
        return None
    except Exception as e:
        return None

def normalize_pharma_company(client_value):
    if not client_value:
        return None
    client_str = str(client_value).lower().strip()
    for key, normalized in PHARMA_COMPANY_MAP.items():
        if key in client_str:
            return normalized
    return str(client_value)

def extract_target_list_data(file_content, filename):
    try:
        ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''

        with tempfile.NamedTemporaryFile(delete=False, suffix=f'.{ext}') as tmp:
            tmp.write(file_content)
            tmp_path = tmp.name

        try:
            if ext == 'xls':
                import xlrd
                wb = xlrd.open_workbook(tmp_path)
                sheet = wb.sheet_by_index(0)
                headers = [sheet.cell_value(0, col) for col in range(sheet.ncols)]
                if sheet.nrows < 2:
                    return {'error': 'No data rows found'}
                row_data = {}
                for idx, header in enumerate(headers):
                    if header:
                        row_data[header] = sheet.cell_value(1, idx)
            else:
                wb = openpyxl.load_workbook(tmp_path, data_only=True)
                sheet = wb.active
                headers = [cell.value for cell in sheet[1]]
                if sheet.max_row < 2:
                    return {'error': 'No data rows found'}
                row_data = {}
                for idx, header in enumerate(headers):
                    cell_value = sheet.cell(row=2, column=idx+1).value
                    if header:
                        row_data[header] = cell_value
        finally:
            os.unlink(tmp_path)

        raw_client_id = find_column_value(row_data, 'client_id')
        client_id = raw_client_id is not None and str(raw_client_id).strip() != ''

        brand_name = find_column_value(row_data, 'brand_name') or ''
        cmi_placement_id = find_column_value(row_data, 'cmi_placement_id')

        extracted = {
            'cmi_placement_id': str(cmi_placement_id) if cmi_placement_id else None,
            'client_placement_id': find_column_value(row_data, 'client_placement_id') or '',
            'client_id': client_id,
            'placement_description': find_column_value(row_data, 'placement_description') or '',
            'supplier': find_column_value(row_data, 'supplier') or 'Matrix Medical Communications',
            'brand_name': brand_name,
            'vehicle_name': find_column_value(row_data, 'vehicle_name') or '',
            'target_list_id': find_column_value(row_data, 'target_list_id') or '',
            'campaign_name': find_column_value(row_data, 'campaign_name') or '',
            'contract_number': find_column_value(row_data, 'contract_number') or '',
            'buying_channel': find_column_value(row_data, 'buying_channel') or '',
            'creative_code_from_file': find_column_value(row_data, 'creative_code') or '',
            'media_tactic_id': find_column_value(row_data, 'media_tactic_id') or ''
        }

        return extracted
    except Exception as e:
        return {'error': str(e)}

def _normalize_npi(val):
    if val is None:
        return None
    if isinstance(val, float):
        if val != val:
            return None
        try:
            val = str(int(val))
        except (ValueError, OverflowError):
            return None
    digits = re.sub(r'\D', '', str(val))
    return digits if len(digits) == 10 else None

def _find_npi_col_index(headers):
    candidates = {'npi', 'npinumber', 'npi#', 'npiid', 'nationalproviderid', 'nationalprovideridentifier'}
    for idx, h in enumerate(headers):
        if not h:
            continue
        cleaned = re.sub(r'[\s_#\-]+', '', str(h)).lower()
        if cleaned in candidates:
            return idx
    return None

def extract_npis_from_target_list(file_content, filename):
    try:
        ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''

        with tempfile.NamedTemporaryFile(delete=False, suffix=f'.{ext}') as tmp:
            tmp.write(file_content)
            tmp_path = tmp.name

        npis = set()
        try:
            if ext == 'xls':
                import xlrd
                wb = xlrd.open_workbook(tmp_path)
                sheet = wb.sheet_by_index(0)
                if sheet.nrows < 2:
                    return []
                headers = [sheet.cell_value(0, c) for c in range(sheet.ncols)]
                npi_col = _find_npi_col_index(headers)
                if npi_col is None:
                    return []
                for r in range(1, sheet.nrows):
                    npi = _normalize_npi(sheet.cell_value(r, npi_col))
                    if npi:
                        npis.add(npi)
            else:
                wb = openpyxl.load_workbook(tmp_path, data_only=True, read_only=True)
                sheet = wb.active
                row_iter = sheet.iter_rows(values_only=True)
                try:
                    headers = list(next(row_iter))
                except StopIteration:
                    wb.close()
                    return []
                npi_col = _find_npi_col_index(headers)
                if npi_col is None:
                    wb.close()
                    return []
                for row in row_iter:
                    if npi_col < len(row):
                        npi = _normalize_npi(row[npi_col])
                        if npi:
                            npis.add(npi)
                wb.close()
        finally:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass

        return list(npis)
    except Exception:
        traceback.print_exc()
        return []

def _merge_target_lists(existing, entry):
    if isinstance(existing, str):
        try:
            existing = json.loads(existing)
        except Exception:
            existing = []
    if not isinstance(existing, list):
        existing = []
    cleaned = [e for e in existing if isinstance(e, dict) and e.get('campaign_id') != entry['campaign_id']]
    cleaned.append(entry)
    return cleaned

def _backfill_chunk(model, chunk, entry, max_retries=4):
    for attempt in range(max_retries + 1):
        session = get_session()
        try:
            rows = session.query(model.id, model.target_lists).filter(
                model.npi.in_(chunk)
            ).order_by(model.id).all()
            for row_id, existing in rows:
                merged = _merge_target_lists(existing, entry)
                session.query(model).filter(model.id == row_id).update(
                    {model.target_lists: merged}, synchronize_session=False
                )
            session.commit()
            return len(rows)
        except OperationalError as e:
            try:
                session.rollback()
            except Exception:
                pass
            if 'deadlock' in str(e).lower() and attempt < max_retries:
                time.sleep(0.4 * (2 ** attempt))
                continue
            traceback.print_exc()
            return 0
        except Exception:
            traceback.print_exc()
            try:
                session.rollback()
            except Exception:
                pass
            return 0
        finally:
            try:
                session.close()
            except Exception:
                pass
    return 0

def _attach_target_list_to_profiles(npis, campaign_id, campaign_name, target_list_id):
    if not npis:
        return
    entry = {
        'campaign_id': str(campaign_id),
        'campaign_name': campaign_name or str(campaign_id),
        'target_list_id': target_list_id or None,
        'attached_at': datetime.utcnow().isoformat()
    }
    BATCH = 500
    npi_list = list(npis)
    universal_hits = 0
    user_hits = 0

    for i in range(0, len(npi_list), BATCH):
        chunk = npi_list[i:i + BATCH]
        universal_hits += _backfill_chunk(UniversalProfile, chunk, entry)

    for i in range(0, len(npi_list), BATCH):
        chunk = npi_list[i:i + BATCH]
        user_hits += _backfill_chunk(UserProfile, chunk, entry)

    print(f"[target_lists] campaign={campaign_id} npis={len(npi_list)} matched universal={universal_hits} user={user_hits}")

def extract_month_from_campaign_name(campaign_name):
    months = {
        'january': 'January', 'jan': 'January',
        'february': 'February', 'feb': 'February',
        'march': 'March', 'mar': 'March',
        'april': 'April', 'apr': 'April',
        'may': 'May',
        'june': 'June', 'jun': 'June',
        'july': 'July', 'jul': 'July',
        'august': 'August', 'aug': 'August',
        'september': 'September', 'sep': 'September', 'sept': 'September',
        'october': 'October', 'oct': 'October',
        'november': 'November', 'nov': 'November',
        'december': 'December', 'dec': 'December'
    }
    campaign_lower = campaign_name.lower()
    for abbr, full in months.items():
        if abbr in campaign_lower:
            return full
    return None

def extract_year_from_campaign_name(campaign_name):
    match = re.search(r'20\d{2}', campaign_name)
    if match:
        return match.group(0)
    return str(datetime.now().year)

def extract_creative_code_from_text(text):
    if not text:
        return None
    patterns = [
        r'(US[-_]?\d{5,6})',
        r'(CA[-_]?\d{4,5}[-_]\d)',
        r'([A-Z]{2,4}[-_][A-Z]{2,4}[-_][A-Z]{2,4}[-_]\d+)',
        r'([A-Z]{2,4}[-_][A-Z]{2,4}[-_][A-Z]{2,4}[-_][A-Z]{2,4}[-_]\d+)',
        r'([A-Z]{2,4}[-_][A-Z]{2,4}[-_]\d+)',
        r'([A-Z]{2,}(?:[-_][A-Z]{2,})+[-_]\d+)',
    ]
    for pattern in patterns:
        match = re.search(pattern, str(text), re.IGNORECASE)
        if match:
            result = match.group(1).upper().replace('_', '-')
            return result
    return None

def extract_ad_size_from_text(text):
    if not text:
        return None
    match = re.search(r'(\d+x\d+)', str(text), re.IGNORECASE)
    if match:
        return match.group(1)
    return None

def extract_tags_data(file_content, filename):
    try:
        ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''

        with tempfile.NamedTemporaryFile(delete=False, suffix=f'.{ext}') as tmp:
            tmp.write(file_content)
            tmp_path = tmp.name

        try:
            headers = None
            header_row_idx = None
            rows_data = []

            if ext == 'xls':
                import xlrd
                wb = xlrd.open_workbook(tmp_path)
                sheet = wb.sheet_by_index(0)

                for row_idx in range(min(sheet.nrows, 20)):
                    row_values = [sheet.cell_value(row_idx, col) for col in range(sheet.ncols)]
                    if any('Placement ID' in str(cell) for cell in row_values if cell):
                        headers = row_values
                        header_row_idx = row_idx
                        break

                if headers and header_row_idx is not None:
                    for row_idx in range(header_row_idx + 1, sheet.nrows):
                        row_values = [sheet.cell_value(row_idx, col) for col in range(sheet.ncols)]
                        rows_data.append(row_values)
            else:
                wb = openpyxl.load_workbook(tmp_path, data_only=True)
                sheet = wb.active

                for row_idx in range(1, min(sheet.max_row + 1, 20)):
                    row_values = [sheet.cell(row=row_idx, column=col).value for col in range(1, sheet.max_column + 1)]
                    if any('Placement ID' in str(cell) for cell in row_values if cell):
                        headers = row_values
                        header_row_idx = row_idx
                        break

                if headers and header_row_idx is not None:
                    for row_idx in range(header_row_idx + 1, sheet.max_row + 1):
                        row_values = [sheet.cell(row=row_idx, column=col).value for col in range(1, sheet.max_column + 1)]
                        rows_data.append(row_values)
        finally:
            os.unlink(tmp_path)

        if not headers or header_row_idx is None:
            return {'error': 'Could not find header row with "Placement ID"'}

        placement_id_col = None
        placement_name_col = None

        for idx, header in enumerate(headers):
            header_str = str(header).strip().lower() if header else ''
            if 'placement id' in header_str and 'name' not in header_str and 'external' not in header_str:
                placement_id_col = idx
            elif 'placement name' in header_str:
                placement_name_col = idx

        if placement_id_col is None:
            return {'error': 'Could not find "Placement ID" column'}

        gcm_placement_id_array = []
        gcm_placement_id_description = []

        for row in rows_data:
            placement_id = row[placement_id_col] if placement_id_col < len(row) else None
            placement_name = row[placement_name_col] if placement_name_col and placement_name_col < len(row) else ''

            if placement_id:
                pid_str = str(int(placement_id) if isinstance(placement_id, float) else placement_id)
                pname_str = str(placement_name) if placement_name else ''

                if pid_str not in gcm_placement_id_array:
                    gcm_placement_id_array.append(pid_str)
                    gcm_placement_id_description.append(pname_str)

        if not gcm_placement_id_array:
            return {'error': 'No placement IDs found'}

        creative_codes = []
        for desc in gcm_placement_id_description:
            code = extract_creative_code_from_text(desc)
            if code and code not in creative_codes:
                creative_codes.append(code)

        return {
            'gcm_placement_id_array': gcm_placement_id_array,
            'gcm_placement_id_description': gcm_placement_id_description,
            'creative_codes_found': creative_codes,
            'total_count': len(gcm_placement_id_array)
        }
    except Exception as e:
        return {'error': str(e)}

def extract_creative_code_from_image(filename):
    return extract_creative_code_from_text(filename)

def validate_and_enrich_with_cmi_contracts(extracted_data):
    try:
        session = get_session()
        cmi_placement_id = extracted_data.get('cmi_placement_id')

        if not cmi_placement_id:
            session.close()
            return extracted_data

        contract = session.query(CMIContractValue).filter(
            CMIContractValue.placement_id == str(cmi_placement_id)
        ).first()

        if contract:
            extracted_data['cmi_placement_id'] = contract.placement_id
            extracted_data['contract_number'] = contract.contract_number or extracted_data.get('contract_number', '')
            extracted_data['vehicle_name'] = contract.vehicle or extracted_data.get('vehicle_name', '')
            extracted_data['buy_component_type'] = contract.buy_component_type or extracted_data.get('buy_component_type', '')

            if contract.brand and extracted_data.get('brand_name', '') != contract.brand:
                extracted_data['brand_name'] = contract.brand
            if contract.placement_description and extracted_data.get('placement_description', '') != contract.placement_description:
                extracted_data['placement_description'] = contract.placement_description

            extracted_data['cmi_validated'] = True
        else:
            extracted_data['cmi_validated'] = False

        session.close()
        return extracted_data

    except Exception as e:
        extracted_data['cmi_validated'] = False
        return extracted_data

@campaigns_bp.route('/<campaign_id>/metadata', methods=['POST'])
def upload_campaign_metadata(campaign_id):
    try:
        has_files = 'target_list' in request.files or 'tags' in request.files or 'ad_images' in request.files
        manual_placement_id = request.form.get('cmi_placement_id', '').strip()

        if not has_files and not manual_placement_id:
            return jsonify({
                'status': 'error',
                'message': 'No files or placement ID provided'
            }), 400

        campaign_name = request.form.get('campaign_name', campaign_id)
        uploaded_by = request.form.get('uploaded_by', 'default_user')
        send_date_str = request.form.get('send_date')
        send_date = None
        if send_date_str:
            try:
                send_date = datetime.strptime(send_date_str, '%Y-%m-%d').date()
            except:
                pass

        extracted_data = {}
        tags_data = None
        creative_codes_from_images = []
        ad_count = 0
        target_list_npis = []

        if 'target_list' in request.files:
            file = request.files['target_list']
            if file and allowed_file(file.filename, 'excel'):
                file_content = file.read()
                extracted_data.update(extract_target_list_data(file_content, file.filename))
                target_list_npis = extract_npis_from_target_list(file_content, file.filename)

        if 'ad_images' in request.files:
            files = request.files.getlist('ad_images')
            ad_count = len(files)
            for file in files:
                if file and allowed_file(file.filename, 'image'):
                    creative_code = extract_creative_code_from_image(file.filename)
                    if creative_code and creative_code not in creative_codes_from_images:
                        creative_codes_from_images.append(creative_code)

        if 'tags' in request.files:
            file = request.files['tags']
            if file and allowed_file(file.filename, 'excel'):
                file_content = file.read()
                tags_data = extract_tags_data(file_content, file.filename)

                if tags_data and 'error' not in tags_data:
                    extracted_data['gcm_placement_id_array'] = tags_data.get('gcm_placement_id_array', [])
                    extracted_data['gcm_placement_id_description'] = tags_data.get('gcm_placement_id_description', [])

                    if not creative_codes_from_images and tags_data.get('creative_codes_found'):
                        creative_codes_from_images = tags_data['creative_codes_found']

        final_creative_code = extracted_data.get('creative_code_from_file') or ''
        if not final_creative_code and creative_codes_from_images:
            final_creative_code = creative_codes_from_images[0]

        if manual_placement_id:
            extracted_data['cmi_placement_id'] = manual_placement_id

        extracted_data = validate_and_enrich_with_cmi_contracts(extracted_data)

        gcm_array = extracted_data.get('gcm_placement_id_array', [])
        gcm_desc_array = extracted_data.get('gcm_placement_id_description', [])

        session = get_session()
        existing = session.query(CampaignReportingMetadata).filter_by(campaign_id=campaign_id).first()

        if existing:
            existing.send_date = send_date or existing.send_date
            existing.client_id = extracted_data.get('client_id', False)
            existing.cmi_placement_id = extracted_data.get('cmi_placement_id') or existing.cmi_placement_id
            existing.client_placement_id = extracted_data.get('client_placement_id') or existing.client_placement_id
            existing.placement_description = extracted_data.get('placement_description') or existing.placement_description
            existing.supplier = extracted_data.get('supplier') or existing.supplier
            existing.brand_name = extracted_data.get('brand_name') or existing.brand_name
            existing.vehicle_name = extracted_data.get('vehicle_name') or existing.vehicle_name
            existing.target_list_id = extracted_data.get('target_list_id') or existing.target_list_id
            existing.campaign_name_from_file = extracted_data.get('campaign_name') or existing.campaign_name_from_file
            existing.creative_code = final_creative_code or existing.creative_code
            existing.gcm_placement_id = json.dumps([]) if not gcm_array else existing.gcm_placement_id
            existing.gcm_placement_id_array = json.dumps(gcm_array) if gcm_array else existing.gcm_placement_id_array
            existing.gcm_placement_id_description = json.dumps(gcm_desc_array) if gcm_desc_array else existing.gcm_placement_id_description
            existing.buy_component_type = extracted_data.get('buy_component_type') or existing.buy_component_type
            existing.contract_number = extracted_data.get('contract_number') or existing.contract_number
            existing.media_tactic_id = extracted_data.get('media_tactic_id') or existing.media_tactic_id
            existing.ad_count = ad_count or existing.ad_count
            existing.raw_metadata = json.dumps(extracted_data, default=str)
            existing.uploaded_by = uploaded_by
            existing.updated_at = datetime.utcnow()
        else:
            metadata = CampaignReportingMetadata(
                campaign_id=campaign_id,
                campaign_name=campaign_name,
                send_date=send_date,
                client_id=extracted_data.get('client_id', False),
                cmi_placement_id=extracted_data.get('cmi_placement_id'),
                client_placement_id=extracted_data.get('client_placement_id'),
                placement_description=extracted_data.get('placement_description'),
                supplier=extracted_data.get('supplier'),
                brand_name=extracted_data.get('brand_name'),
                vehicle_name=extracted_data.get('vehicle_name'),
                target_list_id=extracted_data.get('target_list_id'),
                campaign_name_from_file=extracted_data.get('campaign_name'),
                creative_code=final_creative_code,
                gcm_placement_id=json.dumps([]),
                gcm_placement_id_array=json.dumps(gcm_array) if gcm_array else None,
                gcm_placement_id_description=json.dumps(gcm_desc_array) if gcm_desc_array else None,
                buy_component_type=extracted_data.get('buy_component_type'),
                contract_number=extracted_data.get('contract_number'),
                media_tactic_id=extracted_data.get('media_tactic_id'),
                ad_count=ad_count,
                raw_metadata=json.dumps(extracted_data, default=str),
                uploaded_by=uploaded_by
            )
            session.add(metadata)

        session.commit()
        session.close()

        if target_list_npis:
            threading.Thread(
                target=_attach_target_list_to_profiles,
                args=(target_list_npis, campaign_id, campaign_name, extracted_data.get('target_list_id') or ''),
                daemon=True,
            ).start()

        response_data = {
            'status': 'success',
            'message': 'Metadata extracted successfully',
            'campaign_id': campaign_id,
            'cmi_validated': extracted_data.get('cmi_validated', False),
            'target_list_npis_queued': len(target_list_npis),
            'extracted': {
                'client_id': extracted_data.get('client_id', False),
                'cmi_placement_id': extracted_data.get('cmi_placement_id'),
                'client_placement_id': extracted_data.get('client_placement_id'),
                'brand_name': extracted_data.get('brand_name'),
                'vehicle_name': extracted_data.get('vehicle_name'),
                'target_list_id': extracted_data.get('target_list_id'),
                'creative_code': final_creative_code,
                'contract_number': extracted_data.get('contract_number'),
                'placement_description': extracted_data.get('placement_description'),
                'buy_component_type': extracted_data.get('buy_component_type'),
                'media_tactic_id': extracted_data.get('media_tactic_id')
            }
        }

        if tags_data and 'error' not in tags_data:
            response_data['gcm_selection'] = {
                'gcm_placement_id_array': gcm_array,
                'gcm_placement_id_description': gcm_desc_array,
                'total_count': len(gcm_array)
            }

        return jsonify(response_data), 201

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500

@campaigns_bp.route('/<campaign_id>/metadata', methods=['GET'])
def get_campaign_metadata(campaign_id):
    try:
        session = get_session()
        metadata = session.query(CampaignReportingMetadata).filter_by(campaign_id=campaign_id).first()

        if not metadata:
            session.close()
            return jsonify({
                'status': 'error',
                'message': 'Metadata not found'
            }), 404

        gcm_selected = []
        gcm_array = []
        gcm_desc_array = []
        try:
            if metadata.gcm_placement_id:
                gcm_selected = json.loads(metadata.gcm_placement_id)
            if metadata.gcm_placement_id_array:
                gcm_array = json.loads(metadata.gcm_placement_id_array)
            if metadata.gcm_placement_id_description:
                gcm_desc_array = json.loads(metadata.gcm_placement_id_description)
        except:
            pass

        result = {
            'campaign_id': metadata.campaign_id,
            'campaign_name': metadata.campaign_name,
            'client_id': metadata.client_id,
            'cmi_placement_id': metadata.cmi_placement_id,
            'client_placement_id': metadata.client_placement_id,
            'placement_description': metadata.placement_description,
            'supplier': metadata.supplier,
            'brand_name': metadata.brand_name,
            'vehicle_name': metadata.vehicle_name,
            'target_list_id': metadata.target_list_id,
            'campaign_name_from_file': metadata.campaign_name_from_file,
            'creative_code': metadata.creative_code,
            'gcm_placement_id': gcm_selected,
            'gcm_placement_id_array': gcm_array,
            'gcm_placement_id_description': gcm_desc_array,
            'buy_component_type': metadata.buy_component_type,
            'contract_number': metadata.contract_number,
            'media_tactic_id': metadata.media_tactic_id,
            'uploaded_at': metadata.uploaded_at.isoformat(),
            'uploaded_by': metadata.uploaded_by
        }

        session.close()

        return jsonify({
            'status': 'success',
            'metadata': result
        }), 200

    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500

@campaigns_bp.route('/<campaign_id>/gcm-selection', methods=['PUT'])
def update_gcm_selection(campaign_id):
    try:
        data = request.get_json()
        selected_gcm_ids = data.get('gcm_placement_id', [])

        if not isinstance(selected_gcm_ids, list):
            selected_gcm_ids = [selected_gcm_ids] if selected_gcm_ids else []

        session = get_session()
        metadata = session.query(CampaignReportingMetadata).filter_by(campaign_id=campaign_id).first()

        if not metadata:
            session.close()
            return jsonify({
                'status': 'error',
                'message': 'Metadata not found'
            }), 404

        metadata.gcm_placement_id = json.dumps(selected_gcm_ids)
        metadata.updated_at = datetime.utcnow()

        session.commit()
        session.close()

        return jsonify({
            'status': 'success',
            'message': 'GCM selection updated',
            'gcm_placement_id': selected_gcm_ids
        }), 200

    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500

@campaigns_bp.route('/metadata/all', methods=['GET'])
def get_all_campaign_metadata():
    try:
        session = get_session()
        all_metadata = session.query(CampaignReportingMetadata).all()

        results = []
        for metadata in all_metadata:
            gcm_selected = []
            gcm_array = []
            gcm_desc_array = []
            try:
                if metadata.gcm_placement_id:
                    gcm_selected = json.loads(metadata.gcm_placement_id)
                if metadata.gcm_placement_id_array:
                    gcm_array = json.loads(metadata.gcm_placement_id_array)
                if metadata.gcm_placement_id_description:
                    gcm_desc_array = json.loads(metadata.gcm_placement_id_description)
            except:
                pass

            results.append({
                'campaign_id': metadata.campaign_id,
                'campaign_name': metadata.campaign_name,
                'send_date': metadata.send_date.isoformat() if metadata.send_date else None,
                'client_id': metadata.client_id,
                'cmi_placement_id': metadata.cmi_placement_id,
                'client_placement_id': metadata.client_placement_id,
                'placement_description': metadata.placement_description,
                'brand_name': metadata.brand_name,
                'vehicle_name': metadata.vehicle_name,
                'target_list_id': metadata.target_list_id,
                'campaign_name_from_file': metadata.campaign_name_from_file,
                'creative_code': metadata.creative_code,
                'gcm_placement_id': gcm_selected,
                'gcm_placement_id_array': gcm_array,
                'gcm_placement_id_description': gcm_desc_array,
                'buy_component_type': metadata.buy_component_type,
                'contract_number': metadata.contract_number,
                'media_tactic_id': metadata.media_tactic_id,
                'uploaded_at': metadata.uploaded_at.isoformat() if metadata.uploaded_at else None,
                'uploaded_by': metadata.uploaded_by
            })

        session.close()

        return jsonify({
            'status': 'success',
            'metadata': results,
            'count': len(results)
        }), 200

    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500

@campaigns_bp.route('/gcm/upload', methods=['POST'])
def upload_gcm_tags():
    try:
        if 'tags_file' not in request.files:
            return jsonify({'status': 'error', 'message': 'No tags file provided'}), 400

        file = request.files['tags_file']
        brand = request.form.get('brand', '').strip()

        if not file or not file.filename:
            return jsonify({'status': 'error', 'message': 'No file selected'}), 400

        filename = secure_filename(f"gcm_tags_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename}")
        file_path = os.path.join(UPLOAD_FOLDER, filename)
        file.save(file_path)

        ext = file.filename.rsplit('.', 1)[1].lower()

        placements = []

        if ext == 'xls':
            import xlrd
            wb = xlrd.open_workbook(file_path)
            sheet = wb.sheet_by_index(0)

            header_row_idx = None
            headers = []
            for row_idx in range(min(sheet.nrows, 20)):
                row_values = [sheet.cell_value(row_idx, col) for col in range(sheet.ncols)]
                if any('Placement ID' in str(cell) or 'Placement Id' in str(cell) for cell in row_values if cell):
                    headers = row_values
                    header_row_idx = row_idx
                    break

            if header_row_idx is None:
                os.remove(file_path)
                return jsonify({'status': 'error', 'message': 'Could not find header row with Placement ID'}), 400

            col_map = {}
            for idx, header in enumerate(headers):
                header_str = str(header).strip().lower()
                if 'placement id' in header_str and 'name' not in header_str:
                    col_map['placement_id'] = idx
                elif 'placement name' in header_str or 'placement' in header_str and 'name' in header_str:
                    col_map['placement_name'] = idx
                elif 'advertiser id' in header_str:
                    col_map['advertiser_id'] = idx
                elif 'advertiser name' in header_str or 'advertiser' in header_str and 'name' in header_str:
                    col_map['advertiser_name'] = idx
                elif 'campaign id' in header_str and 'name' not in header_str:
                    col_map['campaign_id'] = idx
                elif 'campaign name' in header_str or 'campaign' in header_str and 'name' in header_str:
                    col_map['campaign_name'] = idx
                elif 'site' in header_str:
                    col_map['site'] = idx
                elif 'start' in header_str and 'date' in header_str:
                    col_map['start_date'] = idx
                elif 'end' in header_str and 'date' in header_str:
                    col_map['end_date'] = idx

            for row_idx in range(header_row_idx + 1, sheet.nrows):
                placement_id = sheet.cell_value(row_idx, col_map.get('placement_id', 0)) if 'placement_id' in col_map else None
                if placement_id:
                    placements.append({
                        'gcm_placement_id': str(int(placement_id) if isinstance(placement_id, float) else placement_id),
                        'placement_name': str(sheet.cell_value(row_idx, col_map['placement_name'])) if 'placement_name' in col_map else None,
                        'advertiser_id': str(sheet.cell_value(row_idx, col_map['advertiser_id'])) if 'advertiser_id' in col_map else None,
                        'advertiser_name': str(sheet.cell_value(row_idx, col_map['advertiser_name'])) if 'advertiser_name' in col_map else None,
                        'campaign_id': str(sheet.cell_value(row_idx, col_map['campaign_id'])) if 'campaign_id' in col_map else None,
                        'campaign_name': str(sheet.cell_value(row_idx, col_map['campaign_name'])) if 'campaign_name' in col_map else None,
                        'site': str(sheet.cell_value(row_idx, col_map['site'])) if 'site' in col_map else None,
                        'brand': brand,
                        'source_file': file.filename
                    })

        elif ext == 'xlsx':
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet = wb.active

            header_row_idx = None
            headers = []
            for row_idx in range(1, min(sheet.max_row + 1, 20)):
                row_values = [sheet.cell(row=row_idx, column=col).value for col in range(1, sheet.max_column + 1)]
                if any('Placement ID' in str(cell) or 'Placement Id' in str(cell) for cell in row_values if cell):
                    headers = row_values
                    header_row_idx = row_idx
                    break

            if header_row_idx is None:
                os.remove(file_path)
                return jsonify({'status': 'error', 'message': 'Could not find header row with Placement ID'}), 400

            col_map = {}
            for idx, header in enumerate(headers):
                if header:
                    header_str = str(header).strip().lower()
                    if 'placement id' in header_str and 'name' not in header_str:
                        col_map['placement_id'] = idx + 1
                    elif 'placement name' in header_str:
                        col_map['placement_name'] = idx + 1
                    elif 'advertiser id' in header_str:
                        col_map['advertiser_id'] = idx + 1
                    elif 'advertiser name' in header_str:
                        col_map['advertiser_name'] = idx + 1
                    elif 'campaign id' in header_str and 'name' not in header_str:
                        col_map['campaign_id'] = idx + 1
                    elif 'campaign name' in header_str:
                        col_map['campaign_name'] = idx + 1
                    elif 'site' in header_str:
                        col_map['site'] = idx + 1
                    elif 'start' in header_str and 'date' in header_str:
                        col_map['start_date'] = idx + 1
                    elif 'end' in header_str and 'date' in header_str:
                        col_map['end_date'] = idx + 1

            for row_idx in range(header_row_idx + 1, sheet.max_row + 1):
                placement_id = sheet.cell(row=row_idx, column=col_map.get('placement_id', 1)).value if 'placement_id' in col_map else None
                if placement_id:
                    placements.append({
                        'gcm_placement_id': str(int(placement_id) if isinstance(placement_id, float) else placement_id),
                        'placement_name': str(sheet.cell(row=row_idx, column=col_map['placement_name']).value) if 'placement_name' in col_map else None,
                        'advertiser_id': str(sheet.cell(row=row_idx, column=col_map['advertiser_id']).value) if 'advertiser_id' in col_map else None,
                        'advertiser_name': str(sheet.cell(row=row_idx, column=col_map['advertiser_name']).value) if 'advertiser_name' in col_map else None,
                        'campaign_id': str(sheet.cell(row=row_idx, column=col_map['campaign_id']).value) if 'campaign_id' in col_map else None,
                        'campaign_name': str(sheet.cell(row=row_idx, column=col_map['campaign_name']).value) if 'campaign_name' in col_map else None,
                        'site': str(sheet.cell(row=row_idx, column=col_map['site']).value) if 'site' in col_map else None,
                        'brand': brand,
                        'source_file': file.filename
                    })

        else:
            os.remove(file_path)
            return jsonify({'status': 'error', 'message': 'Unsupported file format. Use .xls or .xlsx'}), 400

        if not placements:
            os.remove(file_path)
            return jsonify({'status': 'error', 'message': 'No placements found in file'}), 400

        session = get_session()

        if brand:
            session.query(GCMPlacementLookup).filter(GCMPlacementLookup.brand == brand).delete()

        for p in placements:
            lookup = GCMPlacementLookup(
                gcm_placement_id=p['gcm_placement_id'],
                placement_name=p.get('placement_name'),
                advertiser_id=p.get('advertiser_id'),
                advertiser_name=p.get('advertiser_name'),
                campaign_id=p.get('campaign_id'),
                campaign_name=p.get('campaign_name'),
                site=p.get('site'),
                brand=p.get('brand'),
                source_file=p.get('source_file')
            )
            session.add(lookup)

        session.commit()
        session.close()

        os.remove(file_path)

        return jsonify({
            'status': 'success',
            'message': f'Uploaded {len(placements)} GCM placements',
            'count': len(placements),
            'brand': brand
        }), 201

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500

@campaigns_bp.route('/gcm/placements', methods=['GET'])
def get_gcm_placements():
    try:
        brand = request.args.get('brand', '').strip()

        session = get_session()

        query = session.query(GCMPlacementLookup)
        if brand:
            query = query.filter(GCMPlacementLookup.brand.ilike(f'%{brand}%'))

        placements = query.all()

        results = []
        for p in placements:
            results.append({
                'id': p.id,
                'gcm_placement_id': p.gcm_placement_id,
                'placement_name': p.placement_name,
                'advertiser_id': p.advertiser_id,
                'advertiser_name': p.advertiser_name,
                'campaign_id': p.campaign_id,
                'campaign_name': p.campaign_name,
                'site': p.site,
                'brand': p.brand,
                'source_file': p.source_file
            })

        session.close()

        return jsonify({
            'status': 'success',
            'placements': results,
            'count': len(results)
        }), 200

    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500


@campaigns_bp.route('/gcm/brands', methods=['GET'])
def get_gcm_brands():
    try:
        session = get_session()

        brands = session.query(GCMPlacementLookup.brand).distinct().all()
        brand_list = [b[0] for b in brands if b[0]]

        session.close()

        return jsonify({
            'status': 'success',
            'brands': sorted(brand_list)
        }), 200

    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500

def _compute_metrics(sent, bounces, unique_opens, total_opens, unique_clicks, total_clicks):
    delivered = max(0, sent - bounces)
    delivery_rate = round((delivered / sent) * 100, 2) if sent > 0 else 0.0
    unique_open_rate = round((unique_opens / delivered) * 100, 2) if delivered > 0 else 0.0
    total_open_rate = round((total_opens / delivered) * 100, 2) if delivered > 0 else 0.0
    unique_click_rate = round((unique_clicks / unique_opens) * 100, 2) if unique_opens > 0 else 0.0
    total_click_rate = round((total_clicks / total_opens) * 100, 2) if total_opens > 0 else 0.0
    return {
        'sent': sent,
        'delivered': delivered,
        'bounces': bounces,
        'unique_opens': unique_opens,
        'total_opens': total_opens,
        'unique_clicks': unique_clicks,
        'total_clicks': total_clicks,
        'delivery_rate': delivery_rate,
        'unique_open_rate': unique_open_rate,
        'total_open_rate': total_open_rate,
        'unique_click_rate': unique_click_rate,
        'total_click_rate': total_click_rate,
    }


def _target_filter_jsonb(name):
    return json.dumps([{'campaign_name': name}])


def _target_filter_jsonb_alt(name):
    return json.dumps([{'campaign_id': name}])


@campaigns_bp.route('/<path:campaign_id>/target-list-info', methods=['GET'])
def target_list_info(campaign_id):
    session = get_session()
    try:
        session.execute(text("SET LOCAL work_mem = '64MB'"))
        session.execute(text("SET LOCAL effective_cache_size = '512MB'"))
        f1 = _target_filter_jsonb(campaign_id)
        f2 = _target_filter_jsonb_alt(campaign_id)
        target_npi_sql = text("""
            SELECT COUNT(DISTINCT npi) AS c FROM (
                SELECT npi FROM universal_profiles
                WHERE target_lists IS NOT NULL AND npi IS NOT NULL AND npi <> ''
                  AND (target_lists::jsonb @> CAST(:f1 AS jsonb) OR target_lists::jsonb @> CAST(:f2 AS jsonb))
                UNION
                SELECT npi FROM user_profiles
                WHERE target_lists IS NOT NULL AND npi IS NOT NULL AND npi <> ''
                  AND (target_lists::jsonb @> CAST(:f1 AS jsonb) OR target_lists::jsonb @> CAST(:f2 AS jsonb))
            ) s
        """)
        target_npi_count = session.execute(target_npi_sql, {'f1': f1, 'f2': f2}).scalar() or 0
        target_email_sql = text("""
            SELECT COUNT(DISTINCT u.npi) AS c
            FROM user_profiles u
            WHERE u.email IS NOT NULL AND u.email <> '' AND u.npi IS NOT NULL AND u.npi <> ''
              AND u.npi IN (
                SELECT npi FROM universal_profiles
                WHERE target_lists IS NOT NULL AND npi IS NOT NULL AND npi <> ''
                  AND (target_lists::jsonb @> CAST(:f1 AS jsonb) OR target_lists::jsonb @> CAST(:f2 AS jsonb))
                UNION
                SELECT npi FROM user_profiles
                WHERE target_lists IS NOT NULL AND npi IS NOT NULL AND npi <> ''
                  AND (target_lists::jsonb @> CAST(:f1 AS jsonb) OR target_lists::jsonb @> CAST(:f2 AS jsonb))
              )
        """)
        target_email_count = session.execute(target_email_sql, {'f1': f1, 'f2': f2}).scalar() or 0
        return jsonify({
            'has_target_list': target_npi_count > 0,
            'target_npi_count': int(target_npi_count),
            'target_email_count': int(target_email_count),
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()

@campaigns_bp.route('/<path:campaign_id>/target-list-breakdown', methods=['GET'])
def target_list_breakdown(campaign_id):
    session = get_session()
    try:
        session.execute(text("SET LOCAL work_mem = '64MB'"))
        session.execute(text("SET LOCAL effective_cache_size = '512MB'"))
        f1 = _target_filter_jsonb(campaign_id)
        f2 = _target_filter_jsonb_alt(campaign_id)
        name_dep = campaign_id + ' - Deployment%'
        id_rows = session.execute(text("""
            SELECT DISTINCT campaign_id FROM campaign_interactions
            WHERE campaign_name = :name OR campaign_id = :name OR campaign_name LIKE :name_dep
        """), {'name': campaign_id, 'name_dep': name_dep}).fetchall()
        campaign_ids = [r[0] for r in id_rows if r[0]]
        if not campaign_ids:
            empty = _compute_metrics(0, 0, 0, 0, 0, 0)
            return jsonify({'campaign_id': campaign_id, 'target_list': empty, 'rest_of_audience': empty})
        sql = text("""
            WITH target_emails AS (
                SELECT DISTINCT LOWER(TRIM(u.email)) AS email
                FROM user_profiles u
                WHERE u.email IS NOT NULL AND u.email <> '' AND u.npi IS NOT NULL AND u.npi <> ''
                  AND u.npi IN (
                    SELECT npi FROM universal_profiles
                    WHERE target_lists IS NOT NULL AND npi IS NOT NULL AND npi <> ''
                      AND (target_lists::jsonb @> CAST(:f1 AS jsonb) OR target_lists::jsonb @> CAST(:f2 AS jsonb))
                    UNION
                    SELECT npi FROM user_profiles
                    WHERE target_lists IS NOT NULL AND npi IS NOT NULL AND npi <> ''
                      AND (target_lists::jsonb @> CAST(:f1 AS jsonb) OR target_lists::jsonb @> CAST(:f2 AS jsonb))
                  )
            ),
            events AS (
                SELECT LOWER(TRIM(email)) AS email, event_type
                FROM campaign_interactions
                WHERE campaign_id = ANY(:cids)
            )
            SELECT
                CASE WHEN e.email IN (SELECT email FROM target_emails) THEN 'target' ELSE 'rest' END AS cohort,
                COUNT(DISTINCT e.email) FILTER (WHERE event_type='sent') AS sent,
                COUNT(*) FILTER (WHERE event_type='bounce') AS bounces,
                COUNT(DISTINCT e.email) FILTER (WHERE event_type='open') AS unique_opens,
                COUNT(*) FILTER (WHERE event_type='open') AS total_opens,
                COUNT(DISTINCT e.email) FILTER (WHERE event_type='click') AS unique_clicks,
                COUNT(*) FILTER (WHERE event_type='click') AS total_clicks
            FROM events e
            GROUP BY cohort
        """)
        rows = session.execute(sql, {'cids': campaign_ids, 'f1': f1, 'f2': f2}).fetchall()
        cohorts = {'target': None, 'rest': None}
        for r in rows:
            cohorts[r[0]] = _compute_metrics(
                sent=int(r[1] or 0),
                bounces=int(r[2] or 0),
                unique_opens=int(r[3] or 0),
                total_opens=int(r[4] or 0),
                unique_clicks=int(r[5] or 0),
                total_clicks=int(r[6] or 0),
            )
        empty = _compute_metrics(0, 0, 0, 0, 0, 0)
        return jsonify({
            'campaign_id': campaign_id,
            'target_list': cohorts['target'] or empty,
            'rest_of_audience': cohorts['rest'] or empty,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()