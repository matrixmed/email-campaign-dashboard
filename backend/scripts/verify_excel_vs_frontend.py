import os
import re
import json
import openpyxl
from collections import defaultdict
from dotenv import load_dotenv

load_dotenv(os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), '.env'))

import psycopg2
from psycopg2.extras import RealDictCursor

BASE = r'C:\Users\AndrewDaly\Desktop\email-campaign-dashboard\backend\Lists\Print Lists'


def norm_str(s):
    if s is None:
        return ''
    if isinstance(s, float) and s.is_integer():
        s = str(int(s))
    return re.sub(r'\s+', ' ', str(s).strip().upper())


def get_npi(row):
    npi = row.get('NPI Number') or row.get('NPI')
    if npi is None:
        return None
    if isinstance(npi, float):
        if npi.is_integer():
            npi = int(npi)
        else:
            return None
    npi = str(npi).strip()
    return npi if re.match(r'^\d{10}$', npi) else None


def load_excel(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(h).strip() if h else '' for h in rows[0]]
    out = []
    for r in rows[1:]:
        if r is None or all(v is None or (isinstance(v, str) and not v.strip()) for v in r):
            continue
        d = {header[i]: r[i] for i in range(min(len(header), len(r)))}
        out.append(d)
    wb.close()
    return out


def diff_list(conn, list_name, excel_path):
    excel_rows = load_excel(excel_path)
    excel_npis = set()
    excel_no_npi = []
    for r in excel_rows:
        npi = get_npi(r)
        if npi:
            excel_npis.add(npi)
        else:
            excel_no_npi.append(r)

    cur = conn.cursor(cursor_factory=RealDictCursor)

    cur.execute("""
        SELECT npi FROM universal_profiles
        WHERE print_lists_subscribed ?| ARRAY[%s] AND npi IS NOT NULL AND npi <> ''
    """, (list_name,))
    up_npis = {r['npi'] for r in cur.fetchall()}

    cur.execute("""
        SELECT npi FROM user_profiles
        WHERE print_lists_subscribed ?| ARRAY[%s] AND npi IS NOT NULL AND npi <> ''
    """, (list_name,))
    u_npis = {r['npi'] for r in cur.fetchall()}

    cur.execute("""
        SELECT npi FROM print_only_contacts
        WHERE print_lists_subscribed ?| ARRAY[%s] AND npi IS NOT NULL AND npi <> ''
    """, (list_name,))
    poc_npis = {r['npi'] for r in cur.fetchall()}

    cur.execute("""
        SELECT id, first_name, last_name, address FROM print_only_contacts
        WHERE print_lists_subscribed ?| ARRAY[%s]
          AND (npi IS NULL OR npi = '')
    """, (list_name,))
    poc_no_npi = cur.fetchall()

    db_all_npis = up_npis | u_npis | poc_npis

    in_excel_not_in_db = excel_npis - db_all_npis
    in_db_not_in_excel = db_all_npis - excel_npis

    cur.close()

    return {
        'list_name': list_name,
        'excel_total_rows': len(excel_rows),
        'excel_unique_npis': len(excel_npis),
        'excel_no_npi_rows': len(excel_no_npi),
        'db_up_npis': len(up_npis),
        'db_u_npis': len(u_npis),
        'db_poc_npis': len(poc_npis),
        'db_all_unique_npis': len(db_all_npis),
        'db_poc_no_npi': len(poc_no_npi),
        'missing_in_db': sorted(in_excel_not_in_db),
        'extra_in_db': sorted(in_db_not_in_excel),
        'excel_no_npi_sample': [
            {'name': f"{r.get('First Name','')} {r.get('Last Name','')}".strip(),
             'addr': r.get('Address 1') or r.get('Address Line 1') or r.get('Address  1'),
             'city': r.get('City'), 'state': r.get('State')}
            for r in excel_no_npi[:5]
        ],
        'poc_no_npi_sample': [dict(r) for r in poc_no_npi[:5]],
    }


def main():
    conn = psycopg2.connect(os.getenv('DATABASE_URL'))
    targets = [
        ('JCAD Print List', 'JCAD Print List.xlsx'),
        ('NP+PA Print List', 'NP+PA Print List.xlsx'),
        ('JCAD Comp List', 'JCAD Comp List.xlsx'),
    ]
    for list_name, filename in targets:
        d = diff_list(conn, list_name, os.path.join(BASE, filename))
        print('=' * 70)
        print(f'{list_name}')
        print('=' * 70)
        print(f"  Excel total rows:           {d['excel_total_rows']}")
        print(f"  Excel unique NPIs:          {d['excel_unique_npis']}")
        print(f"  Excel rows w/o NPI:         {d['excel_no_npi_rows']}")
        print(f"  DB universal_profiles:      {d['db_up_npis']}")
        print(f"  DB user_profiles only:      {d['db_u_npis']}")
        print(f"  DB print_only_contacts:     {d['db_poc_npis']}")
        print(f"  DB total UNIQUE (by NPI):   {d['db_all_unique_npis']}")
        print(f"  DB POC without NPI:         {d['db_poc_no_npi']}")
        print()
        print(f"  GAP — missing in DB:        {len(d['missing_in_db'])}  (people on Excel list, NOT subscribed in DB)")
        print(f"  GAP — extra in DB:          {len(d['extra_in_db'])}  (subscribed in DB, NOT in Excel)")
        if d['missing_in_db']:
            print(f"    sample missing NPIs: {d['missing_in_db'][:10]}")
        if d['extra_in_db']:
            print(f"    sample extra NPIs:   {d['extra_in_db'][:10]}")
        if d['excel_no_npi_sample']:
            print(f"  Excel no-NPI samples: {d['excel_no_npi_sample']}")
        if d['poc_no_npi_sample']:
            print(f"  POC no-NPI samples: {d['poc_no_npi_sample']}")
        print()

    conn.close()


if __name__ == '__main__':
    main()
