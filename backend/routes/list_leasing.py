from flask import Blueprint, request, jsonify
from psycopg2.extras import RealDictCursor
import os
import sys
import json
import secrets
from io import BytesIO
from datetime import datetime, timedelta
from collections import defaultdict

import openpyxl
import pandas as pd

sys.path.append(os.path.dirname(os.path.dirname(__file__)))
from db_pool import get_db_connection

list_leasing_bp = Blueprint('list_leasing', __name__)

DERM_TAXONOMY_PREFIX = '207N'
OWNED_SEGMENT = 'Matrix Owned Emails'
IQVIA_SEGMENT = 'IQVIA HCPs'
HLD_SEGMENT = 'HLD HCPs'

STATUS_OWNED_EXPIRE = 'MMC Owned (let license expire)'
STATUS_OWNED = 'MMC Owned'
STATUS_BOTH = 'IQVIA + HLD Licensed'
STATUS_IQVIA = 'IQVIA Licensed'
STATUS_HLD = 'HLD Licensed'
STATUS_MISSING = 'Missing'

_temp_files = {}
_TEMP_TTL_HOURS = 2

_entity_cache = {'ownership': None, 'index': None, 'expires_at': None}
_ENTITY_CACHE_TTL = timedelta(minutes=15)


def _load_taxonomy_mapping():
    path = os.path.join(
        os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
        'src', 'components', 'listanalysis', 'taxonomyMapping.js'
    )
    try:
        with open(path, 'r', encoding='utf-8') as f:
            content = f.read()
    except (FileNotFoundError, OSError):
        return {}
    import re
    return dict(re.findall(r'"([^"]+)":\s*"([^"]+)"', content))


TAXONOMY_MAPPING = _load_taxonomy_mapping()


def _purge_temp_files():
    now = datetime.utcnow()
    expired = [tok for tok, data in _temp_files.items() if data['expires_at'] < now]
    for tok in expired:
        _temp_files.pop(tok, None)


def _normalize_npi(value):
    if value is None:
        return None
    try:
        if isinstance(value, float):
            if value != value:
                return None
            return str(int(value)).zfill(10) if value > 0 else None
        s = str(value).strip()
        if not s:
            return None
        digits = ''.join(c for c in s if c.isdigit())
        if len(digits) == 10:
            return digits
        if len(digits) == 9:
            return '1' + digits
        return None
    except Exception:
        return None


def classify(is_m, is_i, is_h):
    if is_m and (is_i or is_h):
        return STATUS_OWNED_EXPIRE
    if is_m:
        return STATUS_OWNED
    if is_i and is_h:
        return STATUS_BOTH
    if is_i:
        return STATUS_IQVIA
    if is_h:
        return STATUS_HLD
    return STATUS_MISSING


def _load_ownership_set(cursor, segment_name):
    cursor.execute(
        """
        SELECT DISTINCT npi FROM user_profiles
        WHERE ac_segments::jsonb ?| %s
          AND npi IS NOT NULL AND npi <> ''
        """,
        ([segment_name],),
    )
    return {r[0] for r in cursor.fetchall() if r[0]}


def load_ownership_sets(cursor):
    return (
        _load_ownership_set(cursor, OWNED_SEGMENT),
        _load_ownership_set(cursor, IQVIA_SEGMENT),
        _load_ownership_set(cursor, HLD_SEGMENT),
    )


def load_universe(cursor, universe_spec, specialty_filter=None):
    utype = (universe_spec or {}).get('type', 'derm')
    if utype == 'uploaded':
        token = universe_spec.get('file_token')
        data = _temp_files.get(token)
        if not data:
            return []
        result = [(npi, '') for npi in data['npis']]
        return _apply_specialty_filter(result, specialty_filter)

    if utype == 'all_active':
        cursor.execute(
            """
            SELECT npi, primary_specialty
            FROM universal_profiles
            WHERE is_active = TRUE AND npi IS NOT NULL AND npi <> ''
            """
        )
    else:
        cursor.execute(
            """
            SELECT npi, primary_specialty
            FROM universal_profiles
            WHERE primary_taxonomy_code LIKE %s
              AND is_active = TRUE
              AND npi IS NOT NULL AND npi <> ''
            """,
            (DERM_TAXONOMY_PREFIX + '%',),
        )

    rows = [(r[0], r[1] or '') for r in cursor.fetchall()]
    return _apply_specialty_filter(rows, specialty_filter)


def _apply_specialty_filter(rows, specialty_filter):
    if not specialty_filter:
        return rows
    sf = [s.lower() for s in specialty_filter if s]
    if not sf:
        return rows
    return [(n, s) for (n, s) in rows if s and s.lower() in sf]


def _resolve_brand_for_campaign(cursor, campaign_id, campaign_name, brand_cache):
    if campaign_id:
        cursor.execute(
            "SELECT brand_name FROM campaign_reporting_metadata WHERE campaign_id = %s LIMIT 1",
            (campaign_id,),
        )
        row = cursor.fetchone()
        if row and row[0]:
            return row[0]
    return _extract_brand_from_name(cursor, campaign_name, brand_cache)


def _extract_brand_from_name(cursor, name, brand_cache):
    if not name:
        return None
    if 'list' not in brand_cache:
        cursor.execute("SELECT DISTINCT brand FROM brand_editor_agency WHERE is_active = TRUE AND brand IS NOT NULL")
        brand_cache['list'] = [r[0] for r in cursor.fetchall() if r[0]]
    name_lower = name.lower()
    matches = [b for b in brand_cache['list'] if b and b.lower() in name_lower]
    if not matches:
        return None
    return max(matches, key=len)


def _campaigns_in_window(cursor, months):
    cutoff = datetime.utcnow() - timedelta(days=months * 31) if months else None
    if cutoff:
        cursor.execute(
            """
            SELECT DISTINCT (tl->>'campaign_id') AS campaign_id,
                            (tl->>'campaign_name') AS campaign_name,
                            (tl->>'attached_at') AS attached_at
            FROM universal_profiles up,
                 jsonb_array_elements(COALESCE(up.target_lists::jsonb, '[]'::jsonb)) tl
            WHERE (tl->>'attached_at') IS NOT NULL
              AND (tl->>'attached_at')::timestamp >= %s
            """,
            (cutoff,),
        )
    else:
        cursor.execute(
            """
            SELECT DISTINCT (tl->>'campaign_id') AS campaign_id,
                            (tl->>'campaign_name') AS campaign_name,
                            (tl->>'attached_at') AS attached_at
            FROM universal_profiles up,
                 jsonb_array_elements(COALESCE(up.target_lists::jsonb, '[]'::jsonb)) tl
            """
        )
    return [dict(campaign_id=r[0], campaign_name=r[1], attached_at=r[2]) for r in cursor.fetchall()]


def _npis_for_campaign(cursor, campaign_id):
    cursor.execute(
        """
        SELECT DISTINCT up.npi, up.primary_specialty
        FROM universal_profiles up
        WHERE up.target_lists::jsonb @> %s::jsonb
          AND up.npi IS NOT NULL AND up.npi <> ''
        """,
        (json.dumps([{'campaign_id': campaign_id}]),),
    )
    return [(r[0], r[1] or '') for r in cursor.fetchall()]


def load_brand_lists(cursor, brand_sources, time_window_months, market_filter):
    brand_npis = defaultdict(set)
    brand_specs = defaultdict(dict)
    brand_cache = {}

    industry_filter = None
    if market_filter:
        cursor.execute(
            "SELECT brand FROM brand_editor_agency WHERE is_active = TRUE AND industry = ANY(%s)",
            (market_filter,),
        )
        industry_filter = {r[0].lower() for r in cursor.fetchall() if r[0]}

    if not brand_sources:
        for camp in _campaigns_in_window(cursor, time_window_months):
            brand = _resolve_brand_for_campaign(cursor, camp['campaign_id'], camp['campaign_name'], brand_cache)
            if not brand:
                continue
            if industry_filter and brand.lower() not in industry_filter:
                continue
            for npi, spec in _npis_for_campaign(cursor, camp['campaign_id']):
                brand_npis[brand].add(npi)
                if spec and npi not in brand_specs[brand]:
                    brand_specs[brand][npi] = spec
        return brand_npis, brand_specs

    for src in brand_sources:
        stype = src.get('type')
        if stype == 'campaign':
            cids = src.get('campaign_ids', [])
            brand_override = src.get('brand_label')
            for cid in cids:
                cursor.execute(
                    "SELECT campaign_name FROM campaign_reporting_metadata WHERE campaign_id = %s LIMIT 1",
                    (cid,),
                )
                row = cursor.fetchone()
                campaign_name = row[0] if row else None
                brand = brand_override or _resolve_brand_for_campaign(cursor, cid, campaign_name, brand_cache)
                if not brand:
                    continue
                if industry_filter and brand.lower() not in industry_filter:
                    continue
                for npi, spec in _npis_for_campaign(cursor, cid):
                    brand_npis[brand].add(npi)
                    if spec and npi not in brand_specs[brand]:
                        brand_specs[brand][npi] = spec
        elif stype == 'brand':
            target_brand = src.get('brand')
            if not target_brand:
                continue
            for camp in _campaigns_in_window(cursor, time_window_months):
                brand = _resolve_brand_for_campaign(cursor, camp['campaign_id'], camp['campaign_name'], brand_cache)
                if not brand or brand.lower() != target_brand.lower():
                    continue
                for npi, spec in _npis_for_campaign(cursor, camp['campaign_id']):
                    brand_npis[target_brand].add(npi)
                    if spec and npi not in brand_specs[target_brand]:
                        brand_specs[target_brand][npi] = spec
        elif stype == 'uploaded':
            token = src.get('file_token')
            label = src.get('brand_label') or 'Uploaded'
            data = _temp_files.get(token)
            if not data:
                continue
            for npi in data['npis']:
                brand_npis[label].add(npi)
            if data.get('specs'):
                for npi, sp in data['specs'].items():
                    if npi not in brand_specs[label]:
                        brand_specs[label][npi] = sp

    return brand_npis, brand_specs


def compute_views(master, brand_npis, brand_specs, matrix_npis, iqvia_npis, hld_npis):
    master_npi_set = {n for n, _ in master}
    master_specialty = {n: s for n, s in master}

    def flags(npi):
        return npi in matrix_npis, npi in iqvia_npis, npi in hld_npis

    def specialty_for(npi):
        if npi in master_specialty and master_specialty[npi]:
            return master_specialty[npi]
        for b, specs in brand_specs.items():
            if npi in specs and specs[npi]:
                return specs[npi]
        return ''

    master_rows = []
    counts = defaultdict(int)
    for n, sp in master:
        is_m, is_i, is_h = flags(n)
        status = classify(is_m, is_i, is_h)
        counts[status] += 1
        master_rows.append({
            'npi': n,
            'specialty': sp,
            'mmc_owned': 1 if is_m else None,
            'iqvia_licensed': 1 if is_i else None,
            'hld_licensed': 1 if is_h else None,
            'ownership_status': status,
            'total_reach': 1 if (is_m or is_i or is_h) else None,
        })

    pivot = {}
    for r in master_rows:
        d = pivot.setdefault(r['specialty'] or '', {'mmc': 0, 'iqv': 0, 'hld': 0, 'tr': 0, 'n': 0})
        d['n'] += 1
        if r['mmc_owned'] == 1: d['mmc'] += 1
        if r['iqvia_licensed'] == 1: d['iqv'] += 1
        if r['hld_licensed'] == 1: d['hld'] += 1
        if r['total_reach'] == 1: d['tr'] += 1

    master_match = []
    for spec in sorted(pivot.keys()):
        d = pivot[spec]
        master_match.append({
            'specialty': spec,
            'mmc_owned': d['mmc'],
            'iqvia_licensed': d['iqv'],
            'hld_licensed': d['hld'],
            'total_reach': d['tr'],
            'total': d['n'],
            'reach_pct': (d['tr'] / d['n']) if d['n'] else 0,
        })

    npi_brands = defaultdict(list)
    for b, npis in brand_npis.items():
        for n in npis:
            npi_brands[n].append(b)

    combined_all = []
    for brand in sorted(brand_npis.keys()):
        for n in sorted(brand_npis[brand]):
            combined_all.append({'npi': n, 'specialty': specialty_for(n), 'brand': brand})

    top_hcps = []
    for n, brands in npi_brands.items():
        if n in master_npi_set:
            top_hcps.append({'npi': n, 'list_count': len(brands), 'brands': sorted(brands), 'specialty': master_specialty.get(n, '')})
    top_hcps.sort(key=lambda x: (-x['list_count'], x['npi']))

    combined_unique = []
    for n in sorted(master_npi_set):
        brands = npi_brands.get(n)
        if not brands:
            continue
        is_m, is_i, is_h = flags(n)
        combined_unique.append({
            'npi': n,
            'specialty': master_specialty.get(n, ''),
            'brand': sorted(brands)[0],
            'all_brands': sorted(brands),
            'mmc_owned': 1 if is_m else None,
            'iqvia_licensed': 1 if is_i else None,
            'hld_licensed': 1 if is_h else None,
            'ownership_status': classify(is_m, is_i, is_h),
        })

    match_pivot = {}
    for n in npi_brands:
        if n not in master_npi_set:
            continue
        spec = master_specialty.get(n, '')
        d = match_pivot.setdefault(spec, [0, 0, 0, 0])
        is_m, is_i, is_h = flags(n)
        if is_m: d[0] += 1
        if is_i: d[1] += 1
        if is_h: d[2] += 1
        d[3] += 1
    match_results = [
        {'specialty': s, 'mmc_owned': d[0], 'iqvia_licensed': d[1], 'hld_licensed': d[2], 'total': d[3]}
        for s, d in sorted(match_pivot.items())
    ]

    per_brand = {}
    for brand in sorted(brand_npis.keys()):
        per_brand[brand] = [
            {'npi': n, 'specialty': specialty_for(n)}
            for n in sorted(brand_npis[brand])
        ]

    brand_breakdown = []
    for brand in sorted(brand_npis.keys()):
        npis = brand_npis[brand]
        owned = len(npis & matrix_npis)
        iqv = len(npis & iqvia_npis - matrix_npis)
        hld = len(npis & hld_npis - matrix_npis)
        missing = len(npis - matrix_npis - iqvia_npis - hld_npis)
        let_expire = len(npis & matrix_npis & (iqvia_npis | hld_npis))
        brand_breakdown.append({
            'brand': brand,
            'total_npis': len(npis),
            'mmc_owned': owned,
            'iqvia_licensed_not_owned': iqv,
            'hld_licensed_not_owned': hld,
            'missing': missing,
            'let_expire': let_expire,
        })

    summary = [
        {'status': s, 'count': c}
        for s, c in sorted(counts.items(), key=lambda x: -x[1])
    ]

    return {
        'summary': summary,
        'totals': {
            'master_size': len(master),
            'brands': len(brand_npis),
            'matrix_npis_in_db': len(matrix_npis),
            'iqvia_npis_in_db': len(iqvia_npis),
            'hld_npis_in_db': len(hld_npis),
        },
        'brand_breakdown': brand_breakdown,
        'master_match_results': master_match,
        'top_hcps': top_hcps[:5000],
        'combined_unique': combined_unique,
        'combined_all': combined_all,
        'match_results': match_results,
        'per_brand': per_brand,
    }


@list_leasing_bp.route('/available-sources', methods=['GET'])
def available_sources():
    months = request.args.get('window_months', default=12, type=int)
    cutoff = datetime.utcnow() - timedelta(days=months * 31) if months else None

    with get_db_connection() as conn:
        cursor = conn.cursor()
        if cutoff:
            cursor.execute(
                """
                SELECT (tl->>'campaign_id') AS campaign_id,
                       (tl->>'campaign_name') AS campaign_name,
                       MAX((tl->>'attached_at')) AS attached_at,
                       COUNT(DISTINCT up.npi) AS npi_count
                FROM universal_profiles up,
                     jsonb_array_elements(COALESCE(up.target_lists::jsonb, '[]'::jsonb)) tl
                WHERE (tl->>'attached_at') IS NOT NULL
                  AND (tl->>'attached_at')::timestamp >= %s
                GROUP BY (tl->>'campaign_id'), (tl->>'campaign_name')
                ORDER BY attached_at DESC NULLS LAST
                """,
                (cutoff,),
            )
        else:
            cursor.execute(
                """
                SELECT (tl->>'campaign_id') AS campaign_id,
                       (tl->>'campaign_name') AS campaign_name,
                       MAX((tl->>'attached_at')) AS attached_at,
                       COUNT(DISTINCT up.npi) AS npi_count
                FROM universal_profiles up,
                     jsonb_array_elements(COALESCE(up.target_lists::jsonb, '[]'::jsonb)) tl
                GROUP BY (tl->>'campaign_id'), (tl->>'campaign_name')
                ORDER BY attached_at DESC NULLS LAST
                """
            )
        rows = cursor.fetchall()

        brand_cache = {}
        grouped = defaultdict(list)
        unknown = []
        for cid, cname, attached_at, npi_count in rows:
            brand = _resolve_brand_for_campaign(cursor, cid, cname, brand_cache)
            entry = {
                'campaign_id': cid,
                'campaign_name': cname,
                'attached_at': attached_at,
                'npi_count': npi_count,
                'brand': brand,
            }
            if brand:
                grouped[brand].append(entry)
            else:
                unknown.append(entry)

        cursor.execute("SELECT DISTINCT industry FROM brand_editor_agency WHERE is_active = TRUE AND industry IS NOT NULL ORDER BY industry")
        industries = [r[0] for r in cursor.fetchall() if r[0]]

        cursor.execute("SELECT DISTINCT brand, industry FROM brand_editor_agency WHERE is_active = TRUE AND brand IS NOT NULL")
        brand_to_industry = {r[0]: r[1] for r in cursor.fetchall()}

        cursor.close()

    payload = {
        'brands': [
            {
                'brand': brand,
                'industry': brand_to_industry.get(brand),
                'campaigns': sorted(items, key=lambda x: (x['attached_at'] or ''), reverse=True),
                'campaign_count': len(items),
                'total_npi_count': sum(x['npi_count'] or 0 for x in items),
            }
            for brand, items in sorted(grouped.items())
        ],
        'unknown_brand_campaigns': unknown,
        'industries': industries,
    }
    return jsonify(payload), 200


@list_leasing_bp.route('/upload-temp-file', methods=['POST'])
def upload_temp_file():
    _purge_temp_files()

    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    f = request.files['file']
    if not f or not f.filename:
        return jsonify({'error': 'Empty file'}), 400

    brand_label = request.form.get('brand_label') or ''
    role = request.form.get('role') or 'brand'

    filename = f.filename
    ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''
    content = f.read()

    npis = set()
    specs = {}

    try:
        if ext in ('xlsx', 'xls'):
            wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                header_row = None
                for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                    header_row = [str(h).upper().strip() if h is not None else '' for h in row]
                    break
                if not header_row:
                    continue
                npi_idx = None
                spec_idx = None
                for i, h in enumerate(header_row):
                    if npi_idx is None and 'NPI' in h:
                        npi_idx = i
                    if spec_idx is None and 'SPECIALTY' in h:
                        spec_idx = i
                if npi_idx is None:
                    continue
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if npi_idx >= len(row):
                        continue
                    norm = _normalize_npi(row[npi_idx])
                    if not norm:
                        continue
                    npis.add(norm)
                    if spec_idx is not None and spec_idx < len(row) and row[spec_idx]:
                        specs.setdefault(norm, str(row[spec_idx]))
            wb.close()
        elif ext == 'csv':
            df = pd.read_csv(BytesIO(content), dtype=str)
            df.columns = [str(c).upper().strip() for c in df.columns]
            npi_col = None
            spec_col = None
            for c in df.columns:
                if npi_col is None and 'NPI' in c:
                    npi_col = c
                if spec_col is None and 'SPECIALTY' in c:
                    spec_col = c
            if not npi_col:
                return jsonify({'error': 'No NPI column found'}), 400
            for _, row in df.iterrows():
                norm = _normalize_npi(row[npi_col])
                if not norm:
                    continue
                npis.add(norm)
                if spec_col and row.get(spec_col):
                    specs.setdefault(norm, str(row[spec_col]))
        else:
            return jsonify({'error': f'Unsupported file type: {ext}'}), 400
    except Exception as e:
        return jsonify({'error': f'Failed to parse file: {str(e)}'}), 400

    if not npis:
        return jsonify({'error': 'No valid NPIs found in file'}), 400

    token = secrets.token_urlsafe(16)
    _temp_files[token] = {
        'npis': npis,
        'specs': specs,
        'brand_label': brand_label,
        'role': role,
        'filename': filename,
        'expires_at': datetime.utcnow() + timedelta(hours=_TEMP_TTL_HOURS),
    }

    return jsonify({
        'file_token': token,
        'filename': filename,
        'npi_count': len(npis),
        'brand_label': brand_label,
        'role': role,
        'expires_at': _temp_files[token]['expires_at'].isoformat(),
    }), 200


@list_leasing_bp.route('/temp-files', methods=['GET'])
def list_temp_files():
    _purge_temp_files()
    return jsonify({
        'files': [
            {
                'file_token': tok,
                'filename': data['filename'],
                'brand_label': data['brand_label'],
                'role': data.get('role', 'brand'),
                'npi_count': len(data['npis']),
                'expires_at': data['expires_at'].isoformat(),
            }
            for tok, data in _temp_files.items()
        ]
    }), 200


@list_leasing_bp.route('/temp-files/<token>', methods=['DELETE'])
def delete_temp_file(token):
    _temp_files.pop(token, None)
    return jsonify({'ok': True}), 200


@list_leasing_bp.route('/run-analysis', methods=['POST'])
def run_analysis():
    _purge_temp_files()
    spec = request.get_json(silent=True) or {}
    universe = spec.get('universe') or {'type': 'derm'}
    brand_sources = spec.get('brand_sources') or []
    time_window_months = spec.get('time_window_months', 12)
    specialty_filter = spec.get('specialty_filter')
    market_filter = spec.get('market_filter')

    with get_db_connection() as conn:
        cursor = conn.cursor()
        try:
            matrix_npis, iqvia_npis, hld_npis = load_ownership_sets(cursor)
            master = load_universe(cursor, universe, specialty_filter)
            brand_npis, brand_specs = load_brand_lists(cursor, brand_sources, time_window_months, market_filter)
        finally:
            cursor.close()

    result = compute_views(master, brand_npis, brand_specs, matrix_npis, iqvia_npis, hld_npis)
    result['scope'] = {
        'universe': universe,
        'brand_sources': brand_sources,
        'time_window_months': time_window_months,
        'specialty_filter': specialty_filter,
        'market_filter': market_filter,
        'computed_at': datetime.utcnow().isoformat(),
    }
    return jsonify(result), 200


@list_leasing_bp.route('/decisions', methods=['GET'])
def list_decisions():
    status_filter = request.args.get('status')
    action_filter = request.args.get('action')
    npi_filter = request.args.get('npi')

    query = "SELECT id, npi, action, license_source, brand_context, notes, decided_at, decided_by, status, executed_at FROM leasing_decisions WHERE 1=1"
    params = []
    if status_filter:
        query += " AND status = %s"
        params.append(status_filter)
    if action_filter:
        query += " AND action = %s"
        params.append(action_filter)
    if npi_filter:
        query += " AND npi = %s"
        params.append(npi_filter)
    query += " ORDER BY decided_at DESC LIMIT 5000"

    with get_db_connection() as conn:
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        cursor.execute(query, tuple(params))
        rows = cursor.fetchall()
        cursor.close()

    return jsonify({
        'decisions': [
            {
                **dict(r),
                'decided_at': r['decided_at'].isoformat() if r.get('decided_at') else None,
                'executed_at': r['executed_at'].isoformat() if r.get('executed_at') else None,
            }
            for r in rows
        ]
    }), 200


@list_leasing_bp.route('/decisions', methods=['POST'])
def create_decision():
    body = request.get_json(silent=True) or {}
    items = body.get('decisions') if 'decisions' in body else [body]
    if not items:
        return jsonify({'error': 'No decisions provided'}), 400

    created = []
    with get_db_connection() as conn:
        cursor = conn.cursor()
        for item in items:
            npi = item.get('npi')
            action = item.get('action')
            if not npi or action not in ('let_expire', 'keep', 'convert_to_owned'):
                continue
            cursor.execute(
                """
                INSERT INTO leasing_decisions
                    (npi, action, license_source, brand_context, notes, decided_by, status)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
                RETURNING id, decided_at
                """,
                (
                    npi,
                    action,
                    item.get('license_source'),
                    item.get('brand_context'),
                    item.get('notes'),
                    item.get('decided_by'),
                    item.get('status', 'pending'),
                ),
            )
            row = cursor.fetchone()
            created.append({'id': row[0], 'npi': npi, 'action': action, 'decided_at': row[1].isoformat()})
        conn.commit()
        cursor.close()

    return jsonify({'created': created, 'count': len(created)}), 201


@list_leasing_bp.route('/decisions/<int:decision_id>', methods=['PATCH'])
def update_decision(decision_id):
    body = request.get_json(silent=True) or {}
    fields = []
    params = []
    for field in ('action', 'license_source', 'brand_context', 'notes', 'status', 'decided_by'):
        if field in body:
            fields.append(f"{field} = %s")
            params.append(body[field])
    if body.get('status') == 'executed':
        fields.append("executed_at = NOW()")
    elif body.get('status') and body.get('status') != 'executed':
        fields.append("executed_at = NULL")

    if not fields:
        return jsonify({'error': 'No fields to update'}), 400

    params.append(decision_id)
    query = f"UPDATE leasing_decisions SET {', '.join(fields)} WHERE id = %s"

    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute(query, tuple(params))
        conn.commit()
        cursor.close()

    return jsonify({'ok': True, 'id': decision_id}), 200


@list_leasing_bp.route('/decisions/<int:decision_id>', methods=['DELETE'])
def delete_decision(decision_id):
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM leasing_decisions WHERE id = %s", (decision_id,))
        conn.commit()
        cursor.close()
    return jsonify({'ok': True}), 200


@list_leasing_bp.route('/hcp/<npi>', methods=['GET'])
def hcp_detail(npi):
    with get_db_connection() as conn:
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        cursor.execute(
            """
            SELECT npi, first_name, middle_name, last_name, credential, primary_specialty,
                   primary_taxonomy_code, practice_state, practice_city, organization_name,
                   is_active, target_lists
            FROM universal_profiles
            WHERE npi = %s LIMIT 1
            """,
            (npi,),
        )
        universal = cursor.fetchone()

        cursor.execute(
            """
            SELECT id, email, first_name, last_name, specialty, ac_segments, ac_tags,
                   target_lists, is_active, inactive_reason
            FROM user_profiles
            WHERE npi = %s
            """,
            (npi,),
        )
        users = cursor.fetchall()

        emails = [u['email'] for u in users if u.get('email')]
        membership_events = []
        if emails:
            cursor.execute(
                """
                SELECT email, dimension, name, event, at, source
                FROM ac_membership_events
                WHERE email = ANY(%s)
                ORDER BY at DESC LIMIT 500
                """,
                (emails,),
            )
            membership_events = cursor.fetchall()

        engagement = None
        if emails:
            cursor.execute(
                """
                SELECT campaigns_received, campaigns_opened, campaigns_clicked,
                       unique_open_rate, unique_click_rate, first_campaign_date, last_campaign_date
                FROM user_engagement_summary
                WHERE email = ANY(%s)
                ORDER BY campaigns_received DESC NULLS LAST LIMIT 1
                """,
                (emails,),
            )
            engagement = cursor.fetchone()

        cursor.execute(
            """
            SELECT id, action, license_source, brand_context, notes, decided_at, decided_by,
                   status, executed_at
            FROM leasing_decisions
            WHERE npi = %s
            ORDER BY decided_at DESC
            """,
            (npi,),
        )
        decisions = cursor.fetchall()

        cursor.close()

    def serialize_dt(row, fields):
        out = dict(row)
        for f in fields:
            if out.get(f):
                out[f] = out[f].isoformat()
        return out

    return jsonify({
        'npi': npi,
        'universal_profile': dict(universal) if universal else None,
        'user_profiles': [dict(u) for u in users],
        'membership_events': [serialize_dt(e, ['at']) for e in membership_events],
        'engagement': serialize_dt(engagement, ['first_campaign_date', 'last_campaign_date']) if engagement else None,
        'decisions': [serialize_dt(d, ['decided_at', 'executed_at']) for d in decisions],
    }), 200


BRAND_ALIASES = {
    'mbc medical affairs': 'Lilly mBC',
}


def _canonicalize_brand(raw, brand_list_lower, cache):
    if not raw:
        return None
    if raw in cache:
        return cache[raw]
    raw_lower = raw.lower().strip()
    if raw_lower in BRAND_ALIASES:
        canon = BRAND_ALIASES[raw_lower]
    else:
        matches = [(b, bl) for b, bl in brand_list_lower if bl in raw_lower]
        if matches:
            canon = min(matches, key=lambda x: len(x[1]))[0]
        else:
            canon = raw
    cache[raw] = canon
    return canon


def _build_entity_index(cursor, matrix, iqvia, hld, types=None, market_filter=None):
    types = types or {'brand', 'segment', 'tag', 'digital_list'}
    index = {}

    cursor.execute(
        """
        SELECT up.npi, up.ac_segments, up.ac_tags, up.digital_lists_subscribed, up.specialty
        FROM user_profiles up
        WHERE up.npi IS NOT NULL AND up.npi <> ''
        """
    )
    for npi, segments, tags, digital_lists, specialty in cursor.fetchall():
        if 'segment' in types and segments:
            for s in segments:
                if isinstance(s, str) and s.strip():
                    key = ('segment', s)
                    if key not in index:
                        index[key] = {'npis': set(), 'specs': defaultdict(int), 'meta': {}}
                    index[key]['npis'].add(npi)
                    if specialty:
                        index[key]['specs'][specialty] += 1
        if 'tag' in types and tags:
            for t in tags:
                if isinstance(t, str) and t.strip():
                    key = ('tag', t)
                    if key not in index:
                        index[key] = {'npis': set(), 'specs': defaultdict(int), 'meta': {}}
                    index[key]['npis'].add(npi)
                    if specialty:
                        index[key]['specs'][specialty] += 1
        if 'digital_list' in types and digital_lists:
            for d in digital_lists:
                if isinstance(d, str) and d.strip():
                    key = ('digital_list', d)
                    if key not in index:
                        index[key] = {'npis': set(), 'specs': defaultdict(int), 'meta': {}}
                    index[key]['npis'].add(npi)
                    if specialty:
                        index[key]['specs'][specialty] += 1

    if 'brand' in types:
        cursor.execute(
            "SELECT brand, industry, agency, pharma_company FROM brand_editor_agency WHERE is_active = TRUE AND brand IS NOT NULL"
        )
        brand_meta = {r[0]: {'industry': r[1], 'agency': r[2], 'pharma_company': r[3]} for r in cursor.fetchall()}

        cursor.execute("SELECT campaign_id, brand_name FROM campaign_reporting_metadata WHERE brand_name IS NOT NULL")
        campaign_brand_map = {r[0]: r[1] for r in cursor.fetchall() if r[0]}

        brand_list_lower = [(b, b.lower()) for b in brand_meta.keys()]
        canon_cache = {}

        cursor.execute(
            """
            SELECT up.npi, up.primary_taxonomy_code, tl
            FROM universal_profiles up,
                 jsonb_array_elements(COALESCE(up.target_lists::jsonb, '[]'::jsonb)) tl
            WHERE up.npi IS NOT NULL AND up.npi <> ''
            """
        )
        for npi, taxonomy_code, tl in cursor.fetchall():
            if not isinstance(tl, dict):
                continue
            campaign_id = tl.get('campaign_id')
            campaign_name = tl.get('campaign_name') or ''
            raw_brand = campaign_brand_map.get(campaign_id)
            if not raw_brand and campaign_name:
                name_lower = campaign_name.lower()
                matches = [(b, bl) for b, bl in brand_list_lower if bl in name_lower]
                if matches:
                    raw_brand = max(matches, key=lambda x: len(x[1]))[0]
            if not raw_brand:
                continue
            brand = _canonicalize_brand(raw_brand, brand_list_lower, canon_cache)
            if not brand:
                continue
            if market_filter and brand_meta.get(brand, {}).get('industry') != market_filter:
                continue
            key = ('brand', brand)
            if key not in index:
                index[key] = {'npis': set(), 'specs': defaultdict(int), 'meta': brand_meta.get(brand, {})}
            index[key]['npis'].add(npi)
            spec = TAXONOMY_MAPPING.get(taxonomy_code) if taxonomy_code else None
            if spec:
                index[key]['specs'][spec] += 1

    return index


def _format_entity(entity_type, name, data, matrix, iqvia, hld):
    npis = data['npis']
    owned = len(npis & matrix)
    iqv = len(npis & iqvia - matrix)
    hld_c = len(npis & hld - matrix)
    missing = len(npis - matrix - iqvia - hld)
    top_specs = sorted(data.get('specs', {}).items(), key=lambda x: -x[1])
    meta = data.get('meta', {})
    return {
        'type': entity_type,
        'name': name,
        'size': len(npis),
        'owned': owned,
        'iqvia': iqv,
        'hld': hld_c,
        'missing': missing,
        'owned_pct': (owned / len(npis)) if npis else 0,
        'top_specialty': top_specs[0][0] if top_specs else None,
        'market': meta.get('industry'),
        'agency': meta.get('agency'),
        'pharma_company': meta.get('pharma_company'),
    }


ENTITY_CACHE_VERSION = 5


def _get_cached_entity_data(cursor, force_refresh=False):
    now = datetime.utcnow()
    if (not force_refresh
            and _entity_cache.get('version') == ENTITY_CACHE_VERSION
            and _entity_cache['index'] is not None
            and _entity_cache['expires_at']
            and _entity_cache['expires_at'] > now):
        return _entity_cache['ownership'], _entity_cache['index'], _entity_cache['expires_at']

    matrix, iqvia, hld = load_ownership_sets(cursor)
    index = _build_entity_index(cursor, matrix, iqvia, hld)
    expires_at = now + _ENTITY_CACHE_TTL

    _entity_cache['ownership'] = (matrix, iqvia, hld)
    _entity_cache['index'] = index
    _entity_cache['expires_at'] = expires_at
    _entity_cache['version'] = ENTITY_CACHE_VERSION

    return (matrix, iqvia, hld), index, expires_at


@list_leasing_bp.route('/entities', methods=['GET'])
def list_entities():
    type_filter = request.args.get('type')
    search = (request.args.get('search') or '').lower().strip()
    market_filter = request.args.get('market')
    min_size = request.args.get('min_size', default=0, type=int)
    force_refresh = request.args.get('refresh', '').lower() in ('1', 'true', 'yes')

    with get_db_connection() as conn:
        cursor = conn.cursor()
        try:
            (matrix, iqvia, hld), index, expires_at = _get_cached_entity_data(cursor, force_refresh)
        finally:
            cursor.close()

    entities = [
        _format_entity(et, name, data, matrix, iqvia, hld)
        for (et, name), data in index.items()
    ]

    type_counts_all = {
        t: sum(1 for e in entities if e['type'] == t)
        for t in ('brand', 'segment', 'tag', 'digital_list')
    }
    markets_all = sorted({e['market'] for e in entities if e.get('market')})

    if type_filter:
        entities = [e for e in entities if e['type'] == type_filter]
    if market_filter:
        entities = [e for e in entities if e.get('market') == market_filter]
    if search:
        entities = [e for e in entities if search in (e['name'] or '').lower()]
    if min_size:
        entities = [e for e in entities if e['size'] >= min_size]

    entities.sort(key=lambda e: -e['size'])

    return jsonify({
        'entities': entities,
        'count': len(entities),
        'markets': markets_all,
        'type_counts': type_counts_all,
        'cache_expires_at': expires_at.isoformat() if expires_at else None,
    }), 200


@list_leasing_bp.route('/entities/<entity_type>/<path:entity_name>', methods=['GET'])
def entity_detail(entity_type, entity_name):
    overlap_limit = request.args.get('overlap_limit', default=8, type=int)
    sample_limit = request.args.get('sample_limit', default=50, type=int)

    with get_db_connection() as conn:
        cursor = conn.cursor()
        try:
            (matrix, iqvia, hld), index, _expires = _get_cached_entity_data(cursor)
            target = index.get((entity_type, entity_name))
            if not target:
                return jsonify({'error': 'Entity not found'}), 404

            target_npis = target['npis']
            sample_npis = list(target_npis)[:5000]
            sample = []
            if sample_npis:
                cursor.execute(
                    """
                    SELECT npi, first_name, last_name, primary_specialty, practice_state, practice_city, is_active
                    FROM universal_profiles
                    WHERE npi = ANY(%s)
                    ORDER BY last_name NULLS LAST, first_name NULLS LAST
                    LIMIT %s
                    """,
                    (sample_npis, sample_limit),
                )
                for r in cursor.fetchall():
                    sample.append({
                        'npi': r[0],
                        'name': ' '.join(p for p in [r[1], r[2]] if p),
                        'specialty': r[3],
                        'state': r[4],
                        'city': r[5],
                        'is_active': r[6],
                    })
        finally:
            cursor.close()

    overlaps_by_type = defaultdict(list)
    for (other_type, other_name), data in index.items():
        if (other_type, other_name) == (entity_type, entity_name):
            continue
        inter = len(target_npis & data['npis'])
        if inter == 0:
            continue
        overlaps_by_type[other_type].append({
            'type': other_type,
            'name': other_name,
            'overlap': inter,
            'their_size': len(data['npis']),
            'pct_of_target': (inter / len(target_npis)) if target_npis else 0,
        })

    overlaps = {}
    for t, items in overlaps_by_type.items():
        items.sort(key=lambda x: -x['overlap'])
        overlaps[t] = items[:overlap_limit]

    summary = _format_entity(entity_type, entity_name, target, matrix, iqvia, hld)

    return jsonify({
        'entity': summary,
        'top_specialties': [
            {'specialty': s, 'count': c}
            for s, c in sorted(target.get('specs', {}).items(), key=lambda x: -x[1])[:10]
        ],
        'sample_members': sample,
        'overlaps': overlaps,
    }), 200


@list_leasing_bp.route('/redundancy-check', methods=['POST'])
def redundancy_check():
    body = request.get_json(silent=True) or {}
    npis = body.get('npis') or []
    if not npis:
        return jsonify({'error': 'No NPIs provided'}), 400

    with get_db_connection() as conn:
        cursor = conn.cursor()
        matrix_npis, iqvia_npis, hld_npis = load_ownership_sets(cursor)
        cursor.close()

    input_set = {n for n in (_normalize_npi(x) for x in npis) if n}
    owned = input_set & matrix_npis
    iqvia = input_set & iqvia_npis
    hld = input_set & hld_npis
    owned_and_leased = input_set & matrix_npis & (iqvia_npis | hld_npis)

    return jsonify({
        'total': len(input_set),
        'mmc_owned': len(owned),
        'iqvia_licensed': len(iqvia),
        'hld_licensed': len(hld),
        'owned_and_leased': len(owned_and_leased),
        'missing': len(input_set - matrix_npis - iqvia_npis - hld_npis),
        'redundancy_pct': (len(owned_and_leased) / len(input_set)) if input_set else 0,
    }), 200
