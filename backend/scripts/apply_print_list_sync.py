import os
import sys
import re
import json
import openpyxl
from collections import defaultdict
from dotenv import load_dotenv

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
load_dotenv(os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), '.env'))

import psycopg2
from psycopg2.extras import RealDictCursor

BASE = r'C:\Users\AndrewDaly\Desktop\email-campaign-dashboard\backend\Lists\Print Lists'
HT_BASE = os.path.join(BASE, 'Hot Topics List Changes')

DRY_RUN = '--apply' not in sys.argv

def norm_str(s):
    if s is None:
        return ''
    if isinstance(s, float):
        if s.is_integer():
            s = str(int(s))
        else:
            s = str(s)
    s = str(s).strip().upper()
    s = re.sub(r'\s+', ' ', s)
    return s

def norm_zip(s):
    if s is None:
        return ''
    if isinstance(s, float):
        if s.is_integer():
            s = str(int(s))
        else:
            s = str(s)
    s = re.sub(r'\D', '', str(s))
    if len(s) > 5:
        s = s[:5]
    return s.zfill(5) if s else ''

def norm_addr(s):
    if not s:
        return ''
    s = norm_str(s)
    repl = {
        'STREET': 'ST', 'AVENUE': 'AVE', 'BOULEVARD': 'BLVD', 'DRIVE': 'DR',
        'LANE': 'LN', 'ROAD': 'RD', 'COURT': 'CT', 'PLACE': 'PL',
        'SUITE': 'STE', 'APARTMENT': 'APT', 'BUILDING': 'BLDG',
        'NORTH': 'N', 'SOUTH': 'S', 'EAST': 'E', 'WEST': 'W',
        'HIGHWAY': 'HWY', 'PARKWAY': 'PKWY', 'CIRCLE': 'CIR',
    }
    for full, abbr in repl.items():
        s = re.sub(r'\b' + full + r'\b', abbr, s)
    s = re.sub(r'[.,#]', '', s)
    s = re.sub(r'\s+', ' ', s)
    return s.strip()

def norm_name(s):
    if not s:
        return ''
    s = norm_str(s)
    s = re.sub(r'[^A-Z\s-]', '', s)
    return s.strip()


def get_npi(row):
    npi = row.get('NPI Number') or row.get('NPI') or row.get('npi')
    if npi is None:
        return None
    if isinstance(npi, float):
        if npi.is_integer():
            npi = str(int(npi))
        else:
            return None
    npi = str(npi).strip()
    if re.match(r'^\d{10}$', npi):
        return npi
    return None


def excel_val(row, *keys):
    for k in keys:
        if k in row and row[k] not in (None, ''):
            return row[k]
    return None


def get_addr(row):
    return excel_val(row, 'Address 1', 'Address Line 1', 'Address  1')


def get_addr2(row):
    return excel_val(row, 'Address 2', 'Address Line 2', 'Address  2')


def get_zip(row):
    z = excel_val(row, 'Zip', 'Zip Code', 'ZipCode')
    if z is None:
        return None
    if isinstance(z, float):
        if z.is_integer():
            z = int(z)
    s = re.sub(r'\D', '', str(z))
    if len(s) > 5:
        s = s[:5]
    return s.zfill(5) if s else None


def to_str(v):
    if v is None:
        return None
    if isinstance(v, float):
        if v.is_integer():
            v = int(v)
    s = str(v).strip()
    return s if s else None


def load_xlsx_rows(path, sheet=None):
    wb = openpyxl.load_workbook(path, data_only=True)
    sheets = [sheet] if sheet else wb.sheetnames
    out = []
    for sn in sheets:
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = [str(h).strip() if h else '' for h in rows[0]]
        for r in rows[1:]:
            if r is None or all(v is None or (isinstance(v, str) and not v.strip()) for v in r):
                continue
            d = {header[i]: r[i] for i in range(min(len(header), len(r)))}
            d['__sheet'] = sn
            out.append(d)
    wb.close()
    return out


def excel_key(row):
    npi = get_npi(row)
    if npi:
        return ('npi', npi)
    fn = norm_name(row.get('First Name', ''))
    ln = norm_name(row.get('Last Name', ''))
    addr = norm_addr(get_addr(row) or '')
    city = norm_str(row.get('City', ''))
    state = norm_str(row.get('State', ''))
    zc = norm_zip(get_zip(row) or '')
    return ('nm', fn, ln, addr, city, state, zc)


def db_key(row):
    npi = (row.get('npi') or '').strip()
    if npi and re.match(r'^\d{10}$', npi):
        return ('npi', npi)
    return (
        'nm',
        norm_name(row.get('first_name')),
        norm_name(row.get('last_name')),
        norm_addr(row.get('address_1')),
        norm_str(row.get('city')),
        norm_str(row.get('state')),
        norm_zip(row.get('zipcode')),
    )


def add_token(current, token):
    parts = [x.strip() for x in (current or '').split(',') if x.strip()]
    if token not in parts:
        parts.append(token)
    return ','.join(parts) if parts else None


def remove_token(current, token):
    parts = [x.strip() for x in (current or '').split(',') if x.strip() and x.strip() != token]
    return ','.join(parts) if parts else None


def has_token(current, token):
    parts = [x.strip() for x in (current or '').split(',') if x.strip()]
    return token in parts


def log_activity(cur, npi, action, details):
    cur.execute(
        "INSERT INTO print_list_activity_log (npi, action, details, created_at) VALUES (%s, %s, %s, NOW())",
        (npi or 'N/A', action, details[:500] if details else None),
    )


def fetch_active_for_list(conn, list_name):
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute("""
        SELECT id, npi, first_name, last_name, degree, specialty, type_of_professional,
               company, title, address_1, address_2, city, state, zipcode,
               subscribed_lists, unsubscribed_lists, is_comp, is_subscribed
        FROM print_list_subscribers
        WHERE subscribed_lists LIKE %s AND is_subscribed = TRUE
    """, (f'%{list_name}%',))
    rows = cur.fetchall()
    cur.close()
    return [r for r in rows if has_token(r.get('subscribed_lists'), list_name)]


def lookup_existing_by_npi(cur, npi):
    cur.execute(
        "SELECT * FROM print_list_subscribers WHERE npi = %s LIMIT 1",
        (npi,),
    )
    return cur.fetchone()


def lookup_existing_by_name(cur, first_name, last_name, addr=None):
    fn = norm_name(first_name)
    ln = norm_name(last_name)
    if addr:
        cur.execute(
            """SELECT * FROM print_list_subscribers
               WHERE UPPER(TRIM(first_name)) = %s AND UPPER(TRIM(last_name)) = %s
               AND UPPER(TRIM(COALESCE(address_1,''))) LIKE %s
               ORDER BY id LIMIT 1""",
            (fn, ln, f'%{norm_addr(addr)[:15]}%'),
        )
    else:
        cur.execute(
            "SELECT * FROM print_list_subscribers WHERE UPPER(TRIM(first_name)) = %s AND UPPER(TRIM(last_name)) = %s ORDER BY id LIMIT 1",
            (fn, ln),
        )
    return cur.fetchone()


def sync_print_list(conn, list_name, print_file, comp_file=None, comp_sheet=None, type_of_prof_field='Type of Professional'):
    cur = conn.cursor(cursor_factory=RealDictCursor)

    print_rows = load_xlsx_rows(os.path.join(BASE, print_file))
    for r in print_rows:
        r['__is_comp'] = False
    comp_rows = []
    if comp_file:
        comp_rows = load_xlsx_rows(os.path.join(BASE, comp_file), sheet=comp_sheet)
        for r in comp_rows:
            r['__is_comp'] = True

    excel_rows = print_rows + comp_rows

    db_rows = fetch_active_for_list(conn, list_name)
    db_by_key = {db_key(r): r for r in db_rows}
    db_by_name = defaultdict(list)
    for r in db_rows:
        nk = (norm_name(r.get('first_name')), norm_name(r.get('last_name')))
        db_by_name[nk].append(r)

    stats = {'adds': 0, 'edits': 0, 'flag_flips': 0, 'noop': 0, 'matched_db_ids': set()}

    for er in excel_rows:
        is_comp = er['__is_comp']
        npi = get_npi(er)
        ek = excel_key(er)
        match = db_by_key.get(ek)

        if not match:
            nk = (norm_name(er.get('First Name', '')), norm_name(er.get('Last Name', '')))
            for cand in db_by_name.get(nk, []):
                if cand['id'] in stats['matched_db_ids']:
                    continue
                e_addr = norm_addr(get_addr(er) or '')
                d_addr = norm_addr(cand.get('address_1'))
                if e_addr and d_addr and (e_addr[:10] == d_addr[:10] or norm_str(er.get('City')) == norm_str(cand.get('city'))):
                    match = cand
                    break

        if not match and npi:
            cur.execute("SELECT * FROM print_list_subscribers WHERE npi = %s LIMIT 1", (npi,))
            cand = cur.fetchone()
            if cand:
                match = cand

        first_name = to_str(er.get('First Name'))
        last_name = to_str(er.get('Last Name'))
        addr_1 = to_str(get_addr(er))
        addr_2 = to_str(get_addr2(er))
        city = to_str(er.get('City'))
        state = to_str(er.get('State'))
        if state:
            state = state.upper()[:2]
        zipcode = get_zip(er)
        specialty = to_str(er.get('Specialty'))
        type_pro = to_str(er.get(type_of_prof_field) or er.get('Type of Professional'))
        title = to_str(er.get('Title'))
        company = to_str(er.get('Company'))

        if not match:
            new_lists = add_token('', list_name)
            cur.execute("""
                INSERT INTO print_list_subscribers
                (npi, first_name, last_name, specialty, type_of_professional, company, title,
                 address_1, address_2, city, state, zipcode,
                 subscribed_lists, is_subscribed, is_comp, subscribe_date, source, created_at, updated_at)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, TRUE, %s, NOW(), 'file_sync', NOW(), NOW())
                RETURNING id
            """, (npi, first_name, last_name, specialty, type_pro, company, title,
                  addr_1, addr_2, city, state, zipcode, new_lists, is_comp))
            new_id = cur.fetchone()['id']
            log_activity(cur, npi, 'subscribe', f'File sync ADD: {list_name}{" (comp)" if is_comp else ""} - {first_name} {last_name}')
            stats['adds'] += 1
            continue

        stats['matched_db_ids'].add(match['id'])

        update_fields = []
        update_params = []
        addr_changed = False

        e_addr_n = norm_addr(addr_1 or '')
        d_addr_n = norm_addr(match.get('address_1'))
        if addr_1 and e_addr_n and e_addr_n != d_addr_n:
            addr_changed = True

        if addr_changed:
            update_fields.append(
                """address_history = COALESCE(address_history, '[]'::jsonb) || jsonb_build_array(jsonb_build_object(
                    'address', address_1, 'address_2', address_2, 'city', city,
                    'state', state, 'zipcode', zipcode, 'changed_at', NOW()::text
                ))"""
            )
            update_fields.append("old_address_1 = address_1")
            update_fields.append("old_address_2 = address_2")
            update_fields.append("old_city = city")
            update_fields.append("old_state = state")
            update_fields.append("old_zipcode = zipcode")
            update_fields.append("address_1 = %s")
            update_params.append(addr_1)
            update_fields.append("address_2 = %s")
            update_params.append(addr_2)
            update_fields.append("city = %s")
            update_params.append(city)
            update_fields.append("state = %s")
            update_params.append(state)
            update_fields.append("zipcode = %s")
            update_params.append(zipcode)

        scalar_changes = {}

        def stage(col, e_val, d_val):
            if e_val is None or e_val == '':
                return
            if norm_str(e_val) != norm_str(d_val):
                scalar_changes[col] = e_val

        stage('first_name', first_name, match.get('first_name'))
        stage('last_name', last_name, match.get('last_name'))
        stage('specialty', specialty, match.get('specialty'))
        stage('type_of_professional', type_pro, match.get('type_of_professional'))
        stage('company', company, match.get('company'))
        stage('title', title, match.get('title'))

        for col, val in scalar_changes.items():
            update_fields.append(f"{col} = %s")
            update_params.append(val)

        if npi and not match.get('npi'):
            update_fields.append("npi = %s")
            update_params.append(npi)

        cur_sub = match.get('subscribed_lists') or ''
        if not has_token(cur_sub, list_name):
            new_sub = add_token(cur_sub, list_name)
            update_fields.append("subscribed_lists = %s")
            update_params.append(new_sub)

        if not match.get('is_subscribed'):
            update_fields.append("is_subscribed = TRUE")

        flag_flip = bool(match.get('is_comp')) != bool(is_comp)
        if flag_flip:
            update_fields.append("is_comp = %s")
            update_params.append(is_comp)

        had_changes = bool(update_fields)

        if had_changes:
            update_fields.append("updated_at = NOW()")
            update_params.append(match['id'])
            cur.execute(
                f"UPDATE print_list_subscribers SET {', '.join(update_fields)} WHERE id = %s",
                update_params,
            )
            parts = []
            if addr_changed:
                parts.append(f'addr: {match.get("address_1")} -> {addr_1}')
            if scalar_changes:
                parts.append('fields: ' + ', '.join(f'{k}={v}' for k, v in scalar_changes.items()))
            if flag_flip:
                parts.append(f'is_comp: {match.get("is_comp")} -> {is_comp}')
            log_activity(cur, npi or match.get('npi'), 'file_sync_update', f'{list_name}: ' + '; '.join(parts))
            if addr_changed or scalar_changes:
                stats['edits'] += 1
            elif flag_flip:
                stats['flag_flips'] += 1
        else:
            stats['noop'] += 1

    stats['removes'] = []
    for r in db_rows:
        if r['id'] in stats['matched_db_ids']:
            continue
        stats['removes'].append(r)

    cur.close()
    return stats, db_rows


def apply_removes(conn, list_name, removes, delete_sheet_reasons):
    cur = conn.cursor()
    count = 0
    for r in removes:
        npi_key = (r.get('npi') or '').strip()
        name_key = (norm_name(r.get('first_name')), norm_name(r.get('last_name')))
        reason = (
            delete_sheet_reasons.get(('npi', npi_key)) if npi_key else None
        ) or delete_sheet_reasons.get(name_key) or 'Removed via file sync (not in source-of-truth Excel)'

        cur_sub = r.get('subscribed_lists') or ''
        cur_unsub = r.get('unsubscribed_lists') or ''
        new_sub = remove_token(cur_sub, list_name)
        new_unsub = add_token(cur_unsub, list_name)
        is_still_sub = bool(new_sub)

        cur.execute("""
            UPDATE print_list_subscribers
            SET subscribed_lists = %s, unsubscribed_lists = %s, is_subscribed = %s,
                unsubscribe_reason = %s, unsubscribe_date = NOW(), updated_at = NOW()
            WHERE id = %s
        """, (new_sub, new_unsub, is_still_sub, reason, r['id']))
        log_activity(cur, r.get('npi'), 'unsubscribe',
                     f'{list_name}: {r.get("first_name")} {r.get("last_name")} - {reason}')
        count += 1
    cur.close()
    return count


def load_jcad_delete_reasons():
    path = os.path.join(BASE, 'JCAD List Management.xlsx')
    rows = load_xlsx_rows(path, 'Delete')
    reasons = {}
    for r in rows:
        npi = get_npi(r) or to_str(r.get('NPI'))
        reason = to_str(r.get('Notes')) or 'No reason given in Delete sheet'
        if npi and re.match(r'^\d{10}$', str(npi)):
            reasons[('npi', str(npi))] = reason
        nk = (norm_name(r.get('First Name', '')), norm_name(r.get('Last Name', '')))
        reasons.setdefault(nk, reason)
    return reasons


def main():
    print(f'Mode: {"DRY RUN (preview only)" if DRY_RUN else "APPLY"}')
    conn = psycopg2.connect(os.getenv('DATABASE_URL'))

    try:
        jcad_delete_reasons = load_jcad_delete_reasons()
        print(f'Loaded {len(jcad_delete_reasons)} Delete-sheet reasons')

        print('\n--- JCAD (Print + Comp) ---')
        jcad_stats, _ = sync_print_list(conn, 'JCAD', 'JCAD Print List.xlsx', 'JCAD Comp List.xlsx')
        print(f"  adds={jcad_stats['adds']} edits={jcad_stats['edits']} flag_flips={jcad_stats['flag_flips']} noop={jcad_stats['noop']} pending_removes={len(jcad_stats['removes'])}")

        print('\n--- NPPA (Print only) ---')
        nppa_stats, _ = sync_print_list(conn, 'NPPA', 'NP+PA Print List.xlsx')
        print(f"  adds={nppa_stats['adds']} edits={nppa_stats['edits']} flag_flips={nppa_stats['flag_flips']} noop={nppa_stats['noop']} pending_removes={len(nppa_stats['removes'])}")

        print('\n--- BT (Print + Comp) ---')
        bt_stats, _ = sync_print_list(conn, 'BT', 'BT Print List.xlsx', 'BT Comp List.xlsx')
        print(f"  adds={bt_stats['adds']} edits={bt_stats['edits']} flag_flips={bt_stats['flag_flips']} noop={bt_stats['noop']} pending_removes={len(bt_stats['removes'])}")

        print('\n--- Applying removes ---')
        jcad_removed = apply_removes(conn, 'JCAD', jcad_stats['removes'], jcad_delete_reasons)
        nppa_removed = apply_removes(conn, 'NPPA', nppa_stats['removes'], jcad_delete_reasons)
        bt_removed = apply_removes(conn, 'BT', bt_stats['removes'], jcad_delete_reasons)
        print(f'  JCAD removed: {jcad_removed}')
        print(f'  NPPA removed: {nppa_removed}')
        print(f'  BT removed:   {bt_removed}')

        if DRY_RUN:
            print('\nDRY RUN - rolling back.')
            conn.rollback()
        else:
            conn.commit()
            print('\nAPPLIED - committed.')
    except Exception as e:
        conn.rollback()
        print(f'\nERROR (rolled back): {e}')
        raise
    finally:
        conn.close()


if __name__ == '__main__':
    main()