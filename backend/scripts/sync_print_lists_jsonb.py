import os
import re
import sys
import json
import openpyxl
from dotenv import load_dotenv

load_dotenv(os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), '.env'))

import psycopg2
from psycopg2.extras import RealDictCursor

BASE = r'C:\Users\AndrewDaly\Desktop\email-campaign-dashboard\backend\Lists\Print Lists'

DRY_RUN = '--apply' not in sys.argv


def norm_str(s):
    if s is None:
        return ''
    if isinstance(s, float) and s.is_integer():
        s = str(int(s))
    return re.sub(r'\s+', ' ', str(s).strip().upper())


def norm_addr(s):
    if not s:
        return ''
    s = norm_str(s)
    repl = {
        'STREET': 'ST', 'AVENUE': 'AVE', 'BOULEVARD': 'BLVD', 'DRIVE': 'DR',
        'LANE': 'LN', 'ROAD': 'RD', 'COURT': 'CT', 'PLACE': 'PL',
        'SUITE': 'STE', 'APARTMENT': 'APT', 'BUILDING': 'BLDG',
    }
    for full, abbr in repl.items():
        s = re.sub(r'\b' + full + r'\b', abbr, s)
    s = re.sub(r'[.,#]', '', s)
    s = re.sub(r'\s+', ' ', s)
    return s.strip()


def norm_name(s):
    if not s:
        return ''
    s = norm_str(s)
    return re.sub(r'[^A-Z\s-]', '', s).strip()


def get_npi(row):
    npi = row.get('NPI Number') or row.get('NPI')
    if npi is None:
        return None
    if isinstance(npi, float):
        if npi.is_integer():
            npi = int(npi)
        else:
            return None
    npi = str(npi).strip()
    return npi if re.match(r'^\d{10}$', npi) else None


def to_str(v):
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    s = str(v).strip()
    return s if s else None


def get_addr(r):
    return r.get('Address 1') or r.get('Address Line 1') or r.get('Address  1')


def get_zip(r):
    z = r.get('Zip') or r.get('Zip Code')
    if z is None:
        return None
    if isinstance(z, float) and z.is_integer():
        z = int(z)
    s = re.sub(r'\D', '', str(z))
    if len(s) > 5:
        s = s[:5]
    return s.zfill(5) if s else None


def load_excel(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        wb.close()
        return []
    header = [str(h).strip() if h else '' for h in rows[0]]
    out = []
    for r in rows[1:]:
        if r is None or all(v is None or (isinstance(v, str) and not v.strip()) for v in r):
            continue
        out.append({header[i]: r[i] for i in range(min(len(header), len(r)))})
    wb.close()
    return out


def bulk_add_list_by_npi(cur, table, list_name, npis):
    if not npis:
        return 0
    cur.execute(
        f"""UPDATE {table}
            SET print_lists_subscribed = (
                  CASE
                    WHEN print_lists_subscribed IS NULL OR jsonb_typeof(print_lists_subscribed) <> 'array'
                      THEN to_jsonb(ARRAY[%s::text])
                    WHEN print_lists_subscribed @> to_jsonb(ARRAY[%s::text])
                      THEN print_lists_subscribed
                    ELSE print_lists_subscribed || to_jsonb(ARRAY[%s::text])
                  END),
                print_lists_unsubscribed = COALESCE(
                  (SELECT jsonb_agg(e) FROM jsonb_array_elements(print_lists_unsubscribed) e WHERE e <> to_jsonb(%s::text)),
                  '[]'::jsonb),
                updated_at = NOW()
            WHERE npi = ANY(%s::text[])""",
        (list_name, list_name, list_name, list_name, list(npis)),
    )
    return cur.rowcount


def bulk_remove_list_extras(cur, table, list_name, keep_npis, keep_ids=None):
    keep_npis_list = list(keep_npis)
    keep_ids_list = list(keep_ids or [])
    cur.execute(
        f"""UPDATE {table}
            SET print_lists_subscribed = COALESCE(
                  (SELECT jsonb_agg(e) FROM jsonb_array_elements(print_lists_subscribed) e WHERE e <> to_jsonb(%s::text)),
                  '[]'::jsonb),
                print_lists_unsubscribed = (
                  CASE
                    WHEN print_lists_unsubscribed IS NULL OR jsonb_typeof(print_lists_unsubscribed) <> 'array'
                      THEN to_jsonb(ARRAY[%s::text])
                    WHEN print_lists_unsubscribed @> to_jsonb(ARRAY[%s::text])
                      THEN print_lists_unsubscribed
                    ELSE print_lists_unsubscribed || to_jsonb(ARRAY[%s::text])
                  END),
                updated_at = NOW()
            WHERE print_lists_subscribed ?| ARRAY[%s]
              AND (npi IS NULL OR npi = '' OR NOT (npi = ANY(%s::text[])))
              AND NOT (id = ANY(%s::int[]))""",
        (list_name, list_name, list_name, list_name, list_name, keep_npis_list, keep_ids_list),
    )
    return cur.rowcount


def sync_one_list(conn, list_name, excel_path):
    cur = conn.cursor(cursor_factory=RealDictCursor)
    rows = load_excel(excel_path)

    excel_npis = set()
    no_npi_rows = []
    for r in rows:
        npi = get_npi(r)
        if npi:
            excel_npis.add(npi)
        else:
            no_npi_rows.append(r)

    stats = {
        'excel_rows': len(rows),
        'excel_unique_npis': len(excel_npis),
        'excel_no_npi_rows': len(no_npi_rows),
    }

    cur.execute("SELECT npi FROM universal_profiles WHERE npi = ANY(%s::text[])",
                (list(excel_npis),))
    up_existing_npis = {r['npi'] for r in cur.fetchall()}
    cur.execute("SELECT npi FROM user_profiles WHERE npi = ANY(%s::text[])",
                (list(excel_npis),))
    u_existing_npis = {r['npi'] for r in cur.fetchall()}
    cur.execute("SELECT npi FROM print_only_contacts WHERE npi = ANY(%s::text[])",
                (list(excel_npis),))
    p_existing_npis = {r['npi'] for r in cur.fetchall()}

    stats['up_updated'] = bulk_add_list_by_npi(cur, 'universal_profiles', list_name, up_existing_npis)
    stats['u_updated'] = bulk_add_list_by_npi(cur, 'user_profiles', list_name, u_existing_npis)
    stats['poc_updated_by_npi'] = bulk_add_list_by_npi(cur, 'print_only_contacts', list_name, p_existing_npis)

    npis_missing_everywhere = excel_npis - up_existing_npis - u_existing_npis - p_existing_npis
    stats['poc_created_for_missing_npi'] = 0
    if npis_missing_everywhere:
        for npi in npis_missing_everywhere:
            cur.execute("""
                INSERT INTO print_only_contacts (npi, print_lists_subscribed, print_lists_unsubscribed, is_active, source, created_at, updated_at)
                VALUES (%s, to_jsonb(ARRAY[%s::text]), '[]'::jsonb, TRUE, 'excel_sync', NOW(), NOW())
            """, (npi, list_name))
            stats['poc_created_for_missing_npi'] += 1

    matched_poc_ids = set()
    stats['poc_no_npi_matched'] = 0
    stats['poc_no_npi_created'] = 0
    for r in no_npi_rows:
        first_name = to_str(r.get('First Name'))
        last_name = to_str(r.get('Last Name'))
        addr_1 = to_str(get_addr(r))
        city = to_str(r.get('City'))
        state = to_str(r.get('State'))
        if state:
            state = state.upper()[:2]
        zipcode = get_zip(r)
        specialty = to_str(r.get('Specialty'))
        title = to_str(r.get('Title'))
        company = to_str(r.get('Company'))
        addr_norm = norm_addr(addr_1 or '')
        fn = norm_name(first_name)
        ln = norm_name(last_name)

        match_id = None
        if fn and ln:
            cur.execute("""
                SELECT id, address FROM print_only_contacts
                WHERE UPPER(first_name) = %s AND UPPER(last_name) = %s
                  AND (npi IS NULL OR npi = '')
            """, (fn, ln))
            for c in cur.fetchall():
                c_addr = norm_addr(c.get('address'))
                if (addr_norm and c_addr and (addr_norm == c_addr or addr_norm[:15] == c_addr[:15])) or (not addr_norm and not c_addr):
                    match_id = c['id']
                    break

        if match_id is not None:
            cur.execute(
                """UPDATE print_only_contacts
                   SET print_lists_subscribed = (
                     CASE
                       WHEN print_lists_subscribed IS NULL OR jsonb_typeof(print_lists_subscribed) <> 'array'
                         THEN to_jsonb(ARRAY[%s::text])
                       WHEN print_lists_subscribed @> to_jsonb(ARRAY[%s::text])
                         THEN print_lists_subscribed
                       ELSE print_lists_subscribed || to_jsonb(ARRAY[%s::text])
                     END),
                     print_lists_unsubscribed = COALESCE(
                       (SELECT jsonb_agg(e) FROM jsonb_array_elements(print_lists_unsubscribed) e WHERE e <> to_jsonb(%s::text)),
                       '[]'::jsonb),
                     updated_at = NOW()
                   WHERE id = %s""",
                (list_name, list_name, list_name, list_name, match_id),
            )
            matched_poc_ids.add(match_id)
            stats['poc_no_npi_matched'] += 1
        else:
            cur.execute("""
                INSERT INTO print_only_contacts
                  (first_name, last_name, address, city, state, zipcode,
                   specialty, company, title,
                   print_lists_subscribed, print_lists_unsubscribed,
                   is_active, source, created_at, updated_at)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s,
                        to_jsonb(ARRAY[%s::text]), '[]'::jsonb,
                        TRUE, 'excel_sync', NOW(), NOW())
                RETURNING id
            """, (first_name, last_name, addr_1, city, state, zipcode,
                  specialty, company, title, list_name))
            new_id = cur.fetchone()['id']
            matched_poc_ids.add(new_id)
            stats['poc_no_npi_created'] += 1

    stats['up_extras_removed'] = bulk_remove_list_extras(cur, 'universal_profiles', list_name, excel_npis, [])
    stats['u_extras_removed'] = bulk_remove_list_extras(cur, 'user_profiles', list_name, excel_npis, [])
    stats['poc_extras_removed'] = bulk_remove_list_extras(cur, 'print_only_contacts', list_name, excel_npis, matched_poc_ids)

    cur.execute("""SELECT COUNT(DISTINCT npi) AS c FROM (
        SELECT npi FROM universal_profiles WHERE print_lists_subscribed ?| ARRAY[%s] AND npi IS NOT NULL AND npi <> ''
        UNION SELECT npi FROM user_profiles WHERE print_lists_subscribed ?| ARRAY[%s] AND npi IS NOT NULL AND npi <> ''
        UNION SELECT npi FROM print_only_contacts WHERE print_lists_subscribed ?| ARRAY[%s] AND npi IS NOT NULL AND npi <> ''
    ) x""", (list_name, list_name, list_name))
    stats['final_unique_npis'] = cur.fetchone()['c']

    cur.execute("""SELECT COUNT(*) AS c FROM print_only_contacts
        WHERE print_lists_subscribed ?| ARRAY[%s] AND (npi IS NULL OR npi = '')
    """, (list_name,))
    stats['final_poc_no_npi'] = cur.fetchone()['c']

    cur.close()
    return stats


def main():
    print(f"Mode: {'DRY RUN (rollback)' if DRY_RUN else 'APPLY (commit)'}")
    conn = psycopg2.connect(os.getenv('DATABASE_URL'))
    try:
        targets = [
            ('JCAD Print List', 'JCAD Print List.xlsx'),
            ('NP+PA Print List', 'NP+PA Print List.xlsx'),
            ('JCAD Comp List', 'JCAD Comp List.xlsx'),
        ]
        for list_name, filename in targets:
            print()
            print('=' * 70)
            print(list_name)
            print('=' * 70)
            s = sync_one_list(conn, list_name, os.path.join(BASE, filename))
            print(f"  Excel: {s['excel_rows']} rows  ({s['excel_unique_npis']} unique NPIs, {s['excel_no_npi_rows']} no-NPI)")
            print(f"  UP rows updated: {s['up_updated']}")
            print(f"  U rows updated:  {s['u_updated']}")
            print(f"  POC rows updated (by NPI): {s['poc_updated_by_npi']}")
            print(f"  POC created for NPI not in any table: {s['poc_created_for_missing_npi']}")
            print(f"  POC no-NPI matched: {s['poc_no_npi_matched']}, created: {s['poc_no_npi_created']}")
            print(f"  Extras removed — UP: {s['up_extras_removed']}, U: {s['u_extras_removed']}, POC: {s['poc_extras_removed']}")
            print(f"  AFTER — unique NPIs: {s['final_unique_npis']}, POC no-NPI: {s['final_poc_no_npi']}")
        if DRY_RUN:
            conn.rollback()
            print('\nDRY RUN — rolled back.')
        else:
            conn.commit()
            print('\nAPPLIED — committed.')
    except Exception as e:
        conn.rollback()
        print(f'ERROR (rolled back): {e}')
        raise
    finally:
        conn.close()


if __name__ == '__main__':
    main()
